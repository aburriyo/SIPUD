import os
import glob
import openpyxl
import random
import string
from datetime import datetime
from app import create_app, db
from app.models import User, Supplier, Product, Lot, InboundOrder, Sale, SaleItem, LogisticsRoute

app = create_app()

def normalize_product_name(name):
    if not name:
        return None
    return name.strip().lower()

def generate_sku(name, prefix="GEN"):
    base = name.strip().upper()[:3]
    suffix = ''.join(random.choices(string.ascii_uppercase + string.digits, k=4))
    return f"{prefix}-{base}-{suffix}"

def import_data():
    base_path = '/Users/bchavez/Software Inventario 2026/___documentos'
    cuadratura_path = os.path.join(base_path, 'Cuadratura local.xlsx')
    
    with app.app_context():
        print("Starting data import...")
        
        # 1. Create Default Users
        if not User.query.filter_by(username='admin').first():
            db.session.add(User(username='admin', role='admin'))
            db.session.add(User(username='bodega', role='warehouse'))
            db.session.add(User(username='repartidor', role='driver'))
            db.session.commit()
            print("Created default users.")

        # 2. Create Master Products (Minipromo 1-7)
        products = {}
        for i in range(1, 8):
            p_name = f"minipromo {i}"
            product = Product.query.filter_by(name=p_name).first()
            if not product:
                product = Product(name=p_name, sku=f"MP{i}", base_price=0)
                db.session.add(product)
                db.session.commit()
            products[p_name] = product
        print("Created master products.")

        # 3. Import Inbound Orders (Facturas)
        if os.path.exists(cuadratura_path):
            try:
                wb = openpyxl.load_workbook(cuadratura_path, data_only=True)
                if 'Facturas' in wb.sheetnames:
                    ws = wb['Facturas']
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        proveedor_name = row[0]
                        fecha = row[1]
                        factura_num = row[3]
                        producto_name = row[4]
                        cantidad = row[5]
                        
                        if not proveedor_name: continue

                        # Create Supplier
                        supplier = Supplier.query.filter_by(name=proveedor_name).first()
                        if not supplier:
                            supplier = Supplier(name=proveedor_name)
                            db.session.add(supplier)
                            db.session.commit()
                        
                        # Create Inbound Order
                        inbound = InboundOrder.query.filter_by(invoice_number=str(factura_num), supplier_id=supplier.id).first()
                        if not inbound:
                            inbound = InboundOrder(
                                supplier_id=supplier.id,
                                invoice_number=str(factura_num),
                                date_received=fecha if isinstance(fecha, datetime) else datetime.now(),
                                status='received'
                            )
                            db.session.add(inbound)
                            db.session.commit()

                        # Create Lot
                        if producto_name:
                            p_norm = normalize_product_name(producto_name)
                            target_product = None
                            
                            # Try to find existing product
                            for p_name, p_obj in products.items():
                                if p_name in p_norm:
                                    target_product = p_obj
                                    break
                            
                            if not target_product:
                                # Create new product dynamically
                                target_product = Product.query.filter_by(name=p_norm).first()
                                if not target_product:
                                    sku = generate_sku(p_norm, "GEN")
                                    target_product = Product(name=p_norm, sku=sku, base_price=0)
                                    db.session.add(target_product)
                                    db.session.commit()

                            lot = Lot(
                                product_id=target_product.id,
                                inbound_order_id=inbound.id,
                                lot_code=f"LOT-{inbound.id}-{target_product.id}",
                                quantity_initial=cantidad if isinstance(cantidad, (int, float)) else 0,
                                quantity_current=cantidad if isinstance(cantidad, (int, float)) else 0,
                                created_at=inbound.date_received
                            )
                            db.session.add(lot)
                    db.session.commit()
                    print("Imported Facturas.")
            except Exception as e:
                db.session.rollback()
                print(f"Error importing Facturas: {e}")

        # 4. Import Sales (Pedidos)
        pedidos_pattern = os.path.join(base_path, 'Pedidos por mes', '**', '*.xlsx')
        files = glob.glob(pedidos_pattern, recursive=True)
        
        for file_path in files:
            try:
                wb = openpyxl.load_workbook(file_path, data_only=True)
                ws = wb.active
                
                rows = list(ws.iter_rows(min_row=3, values_only=True))
                for row in rows:
                    if not row[0]: continue # No customer name
                    
                    customer = row[0]
                    address = row[1]
                    phone = row[2]
                    prod_name = row[3]
                    qty = row[4]
                    price = row[5]
                    
                    # Create Sale
                    sale = Sale(
                        customer_name=customer,
                        address=address,
                        phone=str(phone),
                        status='delivered', 
                        payment_method=row[8] if len(row) > 8 else None
                    )
                    db.session.add(sale)
                    db.session.commit()
                    
                    # Create SaleItem
                    if prod_name:
                        p_norm = normalize_product_name(prod_name)
                        target_product = None
                         # Try to find existing product
                        for p_name, p_obj in products.items():
                            if p_name in p_norm:
                                target_product = p_obj
                                break
                        
                        if not target_product:
                             target_product = Product.query.filter_by(name=p_norm).first()
                             if not target_product:
                                 sku = generate_sku(p_norm, "TEMP")
                                 target_product = Product(name=p_norm, sku=sku, base_price=0)
                                 db.session.add(target_product)
                                 db.session.commit()
                        
                        item = SaleItem(
                            sale_id=sale.id,
                            product_id=target_product.id,
                            quantity=qty if isinstance(qty, (int, float)) else 1,
                            unit_price=price if isinstance(price, (int, float)) else 0
                        )
                        db.session.add(item)
                db.session.commit()
                print(f"Imported {os.path.basename(file_path)}")
            except Exception as e:
                db.session.rollback()
                print(f"Error importing {file_path}: {e}")

if __name__ == '__main__':
    import_data()
