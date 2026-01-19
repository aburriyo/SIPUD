import openpyxl
import os

file_path = "/Users/bchavez/Software Inventario 2026/___documentos/cuadratura puerto distribucion final con ajste.xlsm"

print(f"Inspecting: {file_path}")

try:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    print(f"Sheet names: {wb.sheetnames}")
    
    for sheet_name in wb.sheetnames:
        print(f"\n--- Sheet: {sheet_name} ---")
        ws = wb[sheet_name]
        
        # Print first 5 rows
        for i, row in enumerate(ws.iter_rows(max_row=5, values_only=True)):
            print(row)
            
except Exception as e:
    print(f"Error: {e}")
