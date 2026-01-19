import openpyxl
from datetime import datetime
from app import create_app, db
from app.models import Tenant, Product, Sale, SaleItem, User, Lot, InboundOrder, Supplier

app = create_app()

EXCEL_PATH = "/Users/bchavez/Software Inventario 2026/___documentos/cuadratura puerto distribucion final con ajste.xlsm"

def run_import():
    with app.app_context():
        print("Starting Import for Puerto Distribución...")
        
        # 1. Get or Create Tenant
        tenant = Tenant.query.filter_by(name="Puerto Distribución").first()
        if not tenant:
            tenant = Tenant(name="Puerto Distribución")
            db.session.add(tenant)
            db.session.commit()
            print(f"Created Tenant: {tenant.name}")
        else:
            print(f"Using Tenant: {tenant.name}")
            
        # 2. Ensure Supplier & Inbound Order exists for initial stock
        supplier = Supplier.query.filter_by(name="Inventario Inicial").first()
        if not supplier:
            supplier = Supplier(name="Inventario Inicial", rut="99999999-9", tenant_id=tenant.id)
            db.session.add(supplier)
            db.session.commit()
            
        inbound = InboundOrder.query.filter_by(invoice_number="INIT-001").first()
        if not inbound:
            inbound = InboundOrder(supplier_id=supplier.id, invoice_number="INIT-001", status="received")
            db.session.add(inbound)
            db.session.commit()

        # Load Workbook
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        
        # 3. Import Products & Prices from 'CONTABILIDAD' (Main Boxes)
        print("\nImporting Products from CONTABILIDAD...")
        ws_cont = wb['CONTABILIDAD ']
        product_map = {} # Name -> Product Object
        
        for row in ws_cont.iter_rows(min_row=2, values_only=True):
            name = row[0] # e.g., 'SEMANAL'
            price = row[1]
            if not name or not price: continue
            
            # Normalize name
            name = name.strip().title()
            
            # Create/Update Product
            product = Product.query.filter_by(tenant_id=tenant.id, name=name).first()
            if not product:
                # Generate SKU based on name hash to ensure uniqueness
                sku_hash = abs(hash(name)) % 10000
                sku_generated = f"{name[:3].upper()}-{sku_hash:04d}"
                
                product = Product(
                    tenant_id=tenant.id,
                    name=name,
                    sku=sku_generated,
                    base_price=float(price)
                )
                db.session.add(product)
                db.session.flush() # Get ID
                
                # Add Initial Stock Lot
                lot = Lot(
                    product_id=product.id,
                    inbound_order_id=inbound.id,
                    lot_code=f"INIT-{product.id}",
                    quantity_initial=100,
                    quantity_current=100
                )
                db.session.add(lot)
                print(f"Created Product: {name} (${price})")
            else:
                product.base_price = float(price)
                print(f"Updated Product: {name}")
            
            product_map[name.lower()] = product
            product_map[name.lower().replace("caja ", "")] = product # Alias
        
        db.session.commit()

        # 4. Import Sales from 'ventas'
        print("\nImporting Sales from 'ventas'...")
        ws_ventas = wb['ventas']
        
        # Get Headers (Product Names)
        headers = []
        for row in ws_ventas.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = row
            break
            
        # Identify Product Columns
        product_cols = {} # Index -> Product Name
        for idx, col in enumerate(headers):
            if not col: continue
            col_str = str(col).strip().lower()
            
            if col_str in ['fecha', 'nº de venta', 'columna1', 'total productos', 'total compra', 'metodo pago', 'repartidor', 'valor factura', 'observaciones ']:
                continue
                
            product_cols[idx] = col_str
            
            if col_str not in product_map:
                p_name = str(col).strip()
                prod = Product.query.filter_by(tenant_id=tenant.id, name=p_name).first()
                if not prod:
                    # Generate SKU based on name hash
                    sku_hash = abs(hash(p_name)) % 10000
                    sku_generated = f"{p_name[:3].upper()}-{sku_hash:04d}"
                    
                    prod = Product(
                        tenant_id=tenant.id,
                        name=p_name,
                        sku=sku_generated,
                        base_price=0 
                    )
                    db.session.add(prod)
                    db.session.flush()
                    
                    # Add Initial Stock Lot
                    lot = Lot(
                        product_id=prod.id,
                        inbound_order_id=inbound.id,
                        lot_code=f"INIT-{prod.id}",
                        quantity_initial=100,
                        quantity_current=100
                    )
                    db.session.add(lot)
                    print(f"Created Item Product: {p_name}")
                product_map[col_str] = prod
        
        db.session.commit()

        # Iterate Sales Rows
        sales_count = 0
        for row in ws_ventas.iter_rows(min_row=2, values_only=True):
            date_val = row[0]
            if not date_val: continue
            
            # Create Sale
            sale = Sale(
                tenant_id=tenant.id,
                date_created=date_val if isinstance(date_val, datetime) else datetime.now(),
                customer_name="Cliente Importado", 
                status="delivered",
                payment_method="Desconocido"
            )
            db.session.add(sale)
            db.session.flush()
            
            # Add Items
            for idx, prod_name in product_cols.items():
                qty = row[idx]
                if qty and isinstance(qty, (int, float)) and qty > 0:
                    product = product_map.get(prod_name)
                    if product:
                        item = SaleItem(
                            sale_id=sale.id,
                            product_id=product.id,
                            quantity=int(qty),
                            unit_price=product.base_price
                        )
                        db.session.add(item)
            
            sales_count += 1
            
        db.session.commit()
        print(f"\nImport Completed. Imported {sales_count} sales.")

if __name__ == "__main__":
    run_import()
