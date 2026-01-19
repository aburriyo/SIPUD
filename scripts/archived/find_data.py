import openpyxl
import os

file_path = "/Users/bchavez/Software Inventario 2026/___documentos/cuadratura puerto distribucion final con ajste.xlsm"
search_terms = ["Guillermo Clarke", "Jose Andrade", "Valeria Belmar", "Ana Ojeda"]

print(f"Searching for terms {search_terms} in {os.path.basename(file_path)}")

try:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    
    found = False
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        print(f"\nScanning sheet: {sheet}")
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            for col_idx, cell_value in enumerate(row, 1):
                if cell_value and isinstance(cell_value, str):
                    for term in search_terms:
                        if term.lower() in cell_value.lower():
                            print(f"FOUND '{term}' in Sheet '{sheet}', Row {row_idx}, Col {col_idx} (Value: '{cell_value}')")
                            found = True
                            
    if not found:
        print("Terms not found in any sheet.")

except Exception as e:
    print(f"Error reading file: {e}")
