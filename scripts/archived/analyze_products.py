#!/usr/bin/env python
"""
Analyze products from cuadratura file and compare with database
"""
import sys
sys.path.insert(0, '/Users/bchavez/Software Inventario 2026')

from app import create_app, db
from app.models import Tenant, Product
import openpyxl
from collections import defaultdict

app = create_app()

FILE_PATH = "/Users/bchavez/Software Inventario 2026/___documentos/cuadratura puerto distribucion final con ajste.xlsm"

def analyze_products():
    with app.app_context():
        tenant = Tenant.query.filter_by(name="Puerto Distribución").first()
        if not tenant:
            print("ERROR: Puerto Distribución tenant not found!")
            return
        
        print(f"=== Product Analysis for: {tenant.name} ===\n")
        
        # Load workbook
        wb = openpyxl.load_workbook(FILE_PATH, read_only=True, data_only=True)
        
        # Analyze FLUJO sheet (stock flow)
        print("--- FLUJO Sheet (Stock/Promotional Boxes) ---")
        ws_flujo = wb['FLUJO']
        flujo_products = []
        for i, row in enumerate(ws_flujo.iter_rows(min_row=2, max_row=10, values_only=True)):
            if row[0]:  # tipo de promocion
                flujo_products.append({
                    'name': row[0],
                    'received': row[1],
                    'stock': row[2],
                    'in_process': row[3],
                    'sold': row[4]
                })
                print(f"  {row[0]}: Recibido={row[1]}, Stock={row[2]}, Vendido={row[4]}")
        
        # Analyze CAJAS Y PROMOCIONES sheet (box composition)
        print("\n--- CAJAS Y PROMOCIONES Sheet (Box Contents) ---")
        ws_cajas = wb['CAJAS Y PROMOCIONES ']  # Note: trailing space in sheet name
        
        current_box = None
        box_contents = defaultdict(list)
        
        for row in ws_cajas.iter_rows(min_row=1, values_only=True):
            if row[1] and isinstance(row[1], str):
                # Check if it's a box title (SEMANAL, MENSUAL, etc.)
                box_name = row[1].strip().upper()
                if box_name in ['SEMANAL', 'MENSUAL', 'BOLSILLO FELIZ', 'B.FELIZ', 'NIÑOS']:
                    current_box = box_name
                    if box_name == 'B.FELIZ':
                        current_box = 'BOLSILLO FELIZ'
                    print(f"\n  📦 {current_box}:")
                    continue
                
                # Check if it's a product line
                if row[2] and current_box and row[1] != '#':
                    product_name = row[2]
                    price = row[3]
                    qty = row[4]
                    if product_name and isinstance(product_name, str):
                        box_contents[current_box].append({
                            'product': product_name,
                            'price': price,
                            'quantity': qty
                        })
                        print(f"     - {product_name}: {qty} x ${price:,.0f}" if isinstance(price, (int, float)) else f"     - {product_name}")
        
        # Get current products from database
        print("\n--- Current Products in Database ---")
        db_products = Product.query.filter_by(tenant_id=tenant.id).all()
        print(f"Total products in DB: {len(db_products)}\n")
        
        # Categorize by type
        boxes = []
        individual_products = []
        
        for p in db_products:
            name_lower = p.name.lower()
            if any(keyword in name_lower for keyword in ['caja', 'box', 'semanal', 'mensual', 'bolsillo', 'niños']):
                boxes.append(p)
            else:
                individual_products.append(p)
        
        print(f"📦 Box Products: {len(boxes)}")
        for p in boxes:
            print(f"  - {p.name} (${p.base_price:,}) [Stock: {p.total_stock}]")
        
        print(f"\n📦 Individual Products: {len(individual_products)}")
        for p in individual_products[:20]:  # Show first 20
            print(f"  - {p.name} (${p.base_price:,}) [Stock: {p.total_stock}]")
        if len(individual_products) > 20:
            print(f"  ... and {len(individual_products) - 20} more")
        
        # Suggest categories
        print("\n--- Suggested Product Categories ---")
        categories = {
            'Cajas y Promociones': [],
            'Granos y Cereales': [],
            'Aceites': [],
            'Pastas y Fideos': [],
            'Conservas': [],
            'Lácteos': [],
            'Verduras y Vegetales': [],
            'Otros': []
        }
        
        for p in db_products:
            name = p.name.lower()
            if any(kw in name for kw in ['caja', 'semanal', 'mensual', 'bolsillo', 'niños']):
                categories['Cajas y Promociones'].append(p.name)
            elif any(kw in name for kw in ['arroz', 'poroto', 'lenteja', 'garbanzo', 'azucar', 'azúcar']):
                categories['Granos y Cereales'].append(p.name)
            elif any(kw in name for kw in ['aceite', 'oil']):
                categories['Aceites'].append(p.name)
            elif any(kw in name for kw in ['fideo', 'pasta', 'spaghetti', 'corbatita', 'nicola']):
                categories['Pastas y Fideos'].append(p.name)
            elif any(kw in name for kw in ['atun', 'atún', 'jurel', 'salmon', 'salmón', 'sardina', 'mariscos', 'jibia', 'chorito']):
                categories['Conservas'].append(p.name)
            elif any(kw in name for kw in ['leche', 'mantequilla', 'queso', 'yogurt']):
                categories['Lácteos'].append(p.name)
            elif any(kw in name for kw in ['verdura', 'vegetal', 'tomate']):
                categories['Verduras y Vegetales'].append(p.name)
            else:
                categories['Otros'].append(p.name)
        
        for cat, prods in categories.items():
            if prods:
                print(f"\n{cat} ({len(prods)} productos):")
                for prod in prods[:5]:
                    print(f"  - {prod}")
                if len(prods) > 5:
                    print(f"  ... and {len(prods) - 5} more")

if __name__ == '__main__':
    analyze_products()
