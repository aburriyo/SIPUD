#!/usr/bin/env python
"""
Import cuadratura (reconciliation) data from Excel file
This imports sales and product data from the cuadratura workbook
"""
import sys
sys.path.insert(0, '/Users/bchavez/Software Inventario 2026')

from app import create_app, db
from app.models import Tenant, Sale, SaleItem, Product, User
from datetime import datetime
import openpyxl
import hashlib

app = create_app()

FILE_PATH = "/Users/bchavez/Software Inventario 2026/___documentos/cuadratura puerto distribucion final con ajste.xlsm"

def get_or_create_product(tenant, product_name, price=0):
    """Get existing product or create new one"""
    clean_name = product_name.strip() if product_name else "Sin nombre"
    
    product = Product.query.filter_by(tenant_id=tenant.id, name=clean_name).first()
    if not product:
        # Generate SKU from name hash
        sku = f"SKU-{hashlib.md5(clean_name.encode()).hexdigest()[:8].upper()}"
        product = Product(
            tenant_id=tenant.id,
            name=clean_name,
            sku=sku,
            price=price,
            stock=0
        )
        db.session.add(product)
    return product

def import_ventas_sheet(wb, tenant):
    """Import sales from 'ventas' sheet"""
    print("\n=== Importing Sales from 'ventas' sheet ===")
    ws = wb['ventas']
    
    # Get header row (row 1)
    headers = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    
    # Product columns start after basic fields
    # Based on inspection: fecha, Nº de venta, Columna1, then box types, then individual products
    
    sales_count = 0
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # Skip if no date
            continue
            
        fecha = row[0]
        if not isinstance(fecha, datetime):
            continue
            
        # Create sale
        sale = Sale(
            tenant_id=tenant.id,
            customer_name=f"Cliente Cuadratura #{row[1] or sales_count+1}",
            address="Dirección desde cuadratura",
            phone="",
            status="delivered",
            payment_method="efectivo",
            date_created=fecha
        )
        db.session.add(sale)
        db.session.flush()
        
        # Map products from columns
        # Columns: 0=fecha, 1=Nº, 2=Columna1, 3=semanal, 4=bolsillo feliz, 5=mensual, 6=niños, 7+=products
        product_mapping = {
            3: ("Caja Semanal", 25400),
            4: ("Caja Bolsillo Feliz", 31500),
            5: ("Caja Mensual", 81200),
            6: ("Caja Niños", 33900),
            7: ("Arroz El Pais 1 Kg", 4500),
            8: ("Verduras Minuto Verde 1 Kg", 3500),
            9: ("Azucar Flor De Reyes 1 Kg", 2800),
            10: ("Atun Isadora 100 gr", 2200),
            11: ("Aceite Full Frito 10 Lt", 28000),
            12: ("Fideos gentil 400 gr", 1200),
            13: ("Salmon al natural corcovado 160 gr", 4800),
            14: ("Surtidos de mariscos corcovado 190 gr", 3900),
            15: ("Poroto blanco de reyes 1kg", 3200),
            16: ("Jurel san jose 160gr", 2100),
            17: ("Jibia corcovado", 3600),
            18: ("Choritos costa tenglo", 3400),
        }
        
        for col_idx, (prod_name, price) in product_mapping.items():
            if col_idx < len(row) and row[col_idx]:
                qty = row[col_idx]
                if isinstance(qty, (int, float)) and qty > 0:
                    product = get_or_create_product(tenant, prod_name, price)
                    
                    sale_item = SaleItem(
                        sale_id=sale.id,
                        product_id=product.id,
                        quantity=int(qty),
                        unit_price=price
                    )
                    db.session.add(sale_item)
        
        sales_count += 1
        if sales_count % 10 == 0:
            print(f"  Processed {sales_count} sales...")
    
    print(f"  Total: {sales_count} sales imported from ventas sheet")
    return sales_count

def main():
    with app.app_context():
        # Get Puerto Distribución tenant
        tenant = Tenant.query.filter_by(name="Puerto Distribución").first()
        if not tenant:
            print("ERROR: Puerto Distribución tenant not found!")
            return
        
        print(f"Starting cuadratura import for: {tenant.name}")
        
        # Load workbook
        print(f"Loading: {FILE_PATH}")
        wb = openpyxl.load_workbook(FILE_PATH, read_only=True, data_only=True)
        
        # Import sales from ventas sheet
        sales_imported = import_ventas_sheet(wb, tenant)
        
        # Commit all changes
        db.session.commit()
        print(f"\n✅ Cuadratura import completed!")
        print(f"   - {sales_imported} sales added")

if __name__ == '__main__':
    main()
