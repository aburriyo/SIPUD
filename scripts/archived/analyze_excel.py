import openpyxl
import os

file_path = "/Users/bchavez/Software Inventario 2026/___documentos/Pedidos por mes/Pedidos Septiembre/Pedidos 16-09.xlsx"

print(f"Analyzing: {os.path.basename(file_path)}")

try:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    print(f"Sheet names: {wb.sheetnames}")

    for sheet in wb.sheetnames:
        print(f"\n--- Sheet: {sheet} ---")
        ws = wb[sheet]
        
        # Print first 5 rows
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True)):
            print(f"Row {i+1}: {row}")

except Exception as e:
    print(f"Error reading file: {e}")
