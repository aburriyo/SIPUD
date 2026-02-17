from flask import Blueprint, send_file, g, abort, jsonify, request, render_template
from flask_login import login_required, current_user
from app.models import Sale, SaleItem, Product, Wastage, InboundOrder, Payment
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from datetime import datetime, timedelta
from functools import wraps
from collections import defaultdict

bp = Blueprint("reports", __name__, url_prefix="/reports")


def permission_required(module, action='view'):
    """Decorator to check permissions before accessing a route"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                abort(401)
            if not current_user.has_permission(module, action):
                abort(403)
            return f(*args, **kwargs)
        return decorated_function
    return decorator


@bp.route("/sales/excel")
@login_required
@permission_required('reports', 'export')
def export_sales_excel():
    tenant = g.current_tenant

    # Query sales for current tenant
    sales = Sale.objects(tenant=tenant).order_by('-date_created')

    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"

    # Headers con estilo
    headers = ["ID", "Fecha", "Cliente", "Estado", "Items", "Total", "Método Pago"]
    ws.append(headers)

    # Estilo para headers
    header_fill = PatternFill(
        start_color="4F81BD", end_color="4F81BD", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for sale in sales:
        # Calculate total and format items string
        items_str = []
        total = 0
        for item in sale.items:
            subtotal = item.quantity * float(item.unit_price)
            total += subtotal
            product_name = item.product.name if item.product else 'N/A'
            items_str.append(f"{item.quantity}x {product_name}")

        ws.append(
            [
                str(sale.id),
                sale.date_created.strftime("%Y-%m-%d %H:%M"),
                sale.customer_name,
                sale.status,
                ", ".join(items_str),
                total,
                sale.payment_method,
            ]
        )

    # Auto-adjust column widths (approximation)
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (TypeError, AttributeError):  # cell.value puede ser None
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"ventas_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.route("/warehouse/wastage/excel")
@login_required
@permission_required('reports', 'export')
def export_wastage_excel():
    """Exportar historial de mermas a Excel"""
    tenant = g.current_tenant

    wastages = Wastage.objects(tenant=tenant).order_by('-date_created')

    wb = Workbook()
    ws = wb.active
    ws.title = "Mermas"

    # Headers
    headers = ["ID", "Fecha", "Producto", "SKU", "Cantidad", "Razón", "Notas"]
    ws.append(headers)

    # Estilo para headers
    header_fill = PatternFill(
        start_color="E74C3C", end_color="E74C3C", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for wastage in wastages:
        product_name = wastage.product.name if wastage.product else 'N/A'
        product_sku = wastage.product.sku if wastage.product else 'N/A'
        ws.append(
            [
                str(wastage.id),
                wastage.date_created.strftime("%Y-%m-%d %H:%M"),
                product_name,
                product_sku,
                wastage.quantity,
                wastage.reason,
                wastage.notes or "-",
            ]
        )

    # Auto-adjust columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (TypeError, AttributeError):  # cell.value puede ser None
                pass
        adjusted_width = min(50, max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"mermas_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.route("/warehouse/inventory/excel")
@login_required
@permission_required('reports', 'export')
def export_inventory_excel():
    """Exportar inventario completo a Excel"""
    tenant = g.current_tenant

    products = Product.objects(tenant=tenant).order_by('name')

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    # Headers
    headers = [
        "SKU",
        "Nombre",
        "Categoría",
        "Precio Base",
        "Stock Total",
        "Stock Crítico",
        "Estado",
        "Vencimiento",
    ]
    ws.append(headers)

    # Estilo para headers
    header_fill = PatternFill(
        start_color="27AE60", end_color="27AE60", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for product in products:
        stock_status = (
            "CRÍTICO" if product.total_stock <= product.critical_stock else "OK"
        )
        expiry = (
            product.expiry_date.strftime("%Y-%m-%d") if product.expiry_date else "-"
        )

        ws.append(
            [
                product.sku,
                product.name,
                product.category or "-",
                float(product.base_price) if product.base_price else 0,
                product.total_stock,
                product.critical_stock,
                stock_status,
                expiry,
            ]
        )

        # Color para stock crítico
        if product.total_stock <= product.critical_stock:
            for cell in ws[ws.max_row]:
                cell.fill = PatternFill(
                    start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
                )

    # Auto-adjust columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (TypeError, AttributeError):  # cell.value puede ser None
                pass
        adjusted_width = min(40, max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"inventario_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.route("/warehouse/orders/excel")
@login_required
@permission_required('reports', 'export')
def export_orders_excel():
    """Exportar pedidos a proveedores a Excel"""
    tenant = g.current_tenant

    orders = InboundOrder.objects(tenant=tenant).order_by('-date_received')

    wb = Workbook()
    ws = wb.active
    ws.title = "Pedidos"

    # Headers
    headers = [
        "ID",
        "Proveedor",
        "N° Factura",
        "Estado",
        "Total",
        "Fecha Creación",
        "Fecha Recepción",
        "Notas",
    ]
    ws.append(headers)

    # Estilo para headers
    header_fill = PatternFill(
        start_color="3498DB", end_color="3498DB", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for order in orders:
        ws.append(
            [
                str(order.id),
                order.supplier_name,
                order.invoice_number,
                order.status,
                float(order.total) if order.total else 0,
                order.created_at.strftime("%Y-%m-%d %H:%M")
                if order.created_at
                else "-",
                order.date_received.strftime("%Y-%m-%d %H:%M")
                if order.date_received
                else "-",
                order.notes or "-",
            ]
        )

    # Auto-adjust columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (TypeError, AttributeError):  # cell.value puede ser None
                pass
        adjusted_width = min(40, max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"pedidos_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ============================================
# FLUJO DE CAJA
# ============================================

@bp.route("/cashflow")
@login_required
@permission_required('reports', 'view')
def cashflow_view():
    return render_template("cashflow.html")


@bp.route("/cashflow/api")
@login_required
@permission_required('reports', 'view')
def cashflow_api():
    """API que retorna datos de flujo de caja para un rango de fechas"""
    tenant = g.current_tenant

    # Parsear fechas del request
    date_from_str = request.args.get('from')
    date_to_str = request.args.get('to')
    group_by = request.args.get('group', 'day')  # day, week, month

    now = datetime.utcnow()

    if date_from_str:
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d')
    else:
        date_from = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    if date_to_str:
        date_to = datetime.strptime(date_to_str, '%Y-%m-%d') + timedelta(days=1)
    else:
        date_to = now + timedelta(days=1)

    # --- INGRESOS: Pagos recibidos ---
    payments = Payment.objects(
        tenant=tenant,
        date_created__gte=date_from,
        date_created__lt=date_to
    ).order_by('date_created')

    ingresos_by_period = defaultdict(float)
    ingresos_by_method = defaultdict(float)
    payment_details = []
    total_ingresos = 0

    for p in payments:
        key = _period_key(p.date_created, group_by)
        amount = float(p.amount)
        ingresos_by_period[key] += amount
        ingresos_by_method[p.payment_via] += amount
        total_ingresos += amount
        payment_details.append({
            'date': p.date_created.strftime('%Y-%m-%d %H:%M'),
            'amount': amount,
            'method': p.payment_via,
            'reference': p.payment_reference or '',
            'sale_id': str(p.sale.id) if p.sale else '',
            'customer': p.sale.customer_name if p.sale else ''
        })

    # --- EGRESOS: Órdenes de compra recibidas ---
    orders = InboundOrder.objects(
        tenant=tenant,
        date_received__gte=date_from,
        date_received__lt=date_to,
        status__in=['received', 'paid']
    ).order_by('date_received')

    egresos_by_period = defaultdict(float)
    egresos_by_supplier = defaultdict(float)
    order_details = []
    total_egresos = 0

    for o in orders:
        key = _period_key(o.date_received, group_by)
        amount = float(o.total) if o.total else 0
        egresos_by_period[key] += amount
        egresos_by_supplier[o.supplier_name or 'Sin proveedor'] += amount
        total_egresos += amount
        order_details.append({
            'date': o.date_received.strftime('%Y-%m-%d %H:%M') if o.date_received else '',
            'amount': amount,
            'supplier': o.supplier_name or 'Sin proveedor',
            'invoice': o.invoice_number or '',
            'id': str(o.id)
        })

    # --- Construir serie temporal unificada ---
    all_keys = sorted(set(list(ingresos_by_period.keys()) + list(egresos_by_period.keys())))
    timeline = []
    balance_running = 0
    for key in all_keys:
        ing = ingresos_by_period.get(key, 0)
        egr = egresos_by_period.get(key, 0)
        balance_running += ing - egr
        timeline.append({
            'period': key,
            'ingresos': round(ing),
            'egresos': round(egr),
            'neto': round(ing - egr),
            'balance': round(balance_running)
        })

    # --- Ventas pendientes de cobro ---
    pending_sales = Sale.objects(
        tenant=tenant,
        payment_status__in=['pendiente', 'parcial']
    )
    total_por_cobrar = 0
    for s in pending_sales:
        total_por_cobrar += s.total_amount - s.total_paid

    return jsonify({
        'success': True,
        'summary': {
            'total_ingresos': round(total_ingresos),
            'total_egresos': round(total_egresos),
            'balance_neto': round(total_ingresos - total_egresos),
            'por_cobrar': round(total_por_cobrar),
        },
        'timeline': timeline,
        'ingresos_by_method': {k: round(v) for k, v in sorted(ingresos_by_method.items())},
        'egresos_by_supplier': dict(sorted(egresos_by_supplier.items(), key=lambda x: -x[1])[:10]),
        'details': {
            'payments': payment_details[-50:],
            'orders': order_details[-50:]
        },
        'date_range': {
            'from': date_from.strftime('%Y-%m-%d'),
            'to': (date_to - timedelta(days=1)).strftime('%Y-%m-%d')
        }
    })


def _period_key(dt, group_by):
    """Genera la clave de agrupación según el tipo"""
    if group_by == 'month':
        return dt.strftime('%Y-%m')
    elif group_by == 'week':
        # Inicio de la semana (lunes)
        start = dt - timedelta(days=dt.weekday())
        return start.strftime('%Y-%m-%d')
    else:  # day
        return dt.strftime('%Y-%m-%d')


@bp.route("/cashflow/excel")
@login_required
@permission_required('reports', 'export')
def export_cashflow_excel():
    """Exportar flujo de caja a Excel"""
    tenant = g.current_tenant

    date_from_str = request.args.get('from')
    date_to_str = request.args.get('to')

    now = datetime.utcnow()
    if date_from_str:
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d')
    else:
        date_from = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    if date_to_str:
        date_to = datetime.strptime(date_to_str, '%Y-%m-%d') + timedelta(days=1)
    else:
        date_to = now + timedelta(days=1)

    wb = Workbook()

    # === Hoja 1: Resumen ===
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"

    green_fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
    red_fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
    blue_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    money_fmt = '#,##0'

    ws_resumen.append(["Flujo de Caja — Puerto Distribución"])
    ws_resumen['A1'].font = Font(bold=True, size=14)
    ws_resumen.append([f"Período: {date_from.strftime('%d/%m/%Y')} — {(date_to - timedelta(days=1)).strftime('%d/%m/%Y')}"])
    ws_resumen.append([])

    # Pagos recibidos agrupados por día
    payments = Payment.objects(
        tenant=tenant,
        date_created__gte=date_from,
        date_created__lt=date_to
    ).order_by('date_created')

    daily = defaultdict(lambda: {'ingresos': 0, 'egresos': 0})
    total_in = 0
    for p in payments:
        key = p.date_created.strftime('%Y-%m-%d')
        amt = float(p.amount)
        daily[key]['ingresos'] += amt
        total_in += amt

    orders = InboundOrder.objects(
        tenant=tenant,
        date_received__gte=date_from,
        date_received__lt=date_to,
        status__in=['received', 'paid']
    ).order_by('date_received')

    total_out = 0
    for o in orders:
        if o.date_received:
            key = o.date_received.strftime('%Y-%m-%d')
            amt = float(o.total) if o.total else 0
            daily[key]['egresos'] += amt
            total_out += amt

    # Tabla resumen diario
    headers = ["Fecha", "Ingresos", "Egresos", "Neto", "Balance Acumulado"]
    ws_resumen.append(headers)
    for cell in ws_resumen[4]:
        cell.fill = blue_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    balance = 0
    for key in sorted(daily.keys()):
        d = daily[key]
        neto = d['ingresos'] - d['egresos']
        balance += neto
        ws_resumen.append([key, d['ingresos'], d['egresos'], neto, balance])
        row = ws_resumen.max_row
        for col in [2, 3, 4, 5]:
            ws_resumen.cell(row=row, column=col).number_format = money_fmt
        if neto < 0:
            ws_resumen.cell(row=row, column=4).font = Font(color="E74C3C")

    # Totales
    ws_resumen.append([])
    ws_resumen.append(["TOTAL", total_in, total_out, total_in - total_out, ""])
    row = ws_resumen.max_row
    ws_resumen.cell(row=row, column=1).font = Font(bold=True)
    for col in [2, 3, 4]:
        ws_resumen.cell(row=row, column=col).font = Font(bold=True)
        ws_resumen.cell(row=row, column=col).number_format = money_fmt

    # === Hoja 2: Ingresos (detalle) ===
    ws_ing = wb.create_sheet("Ingresos")
    headers_ing = ["Fecha", "Cliente", "Monto", "Método", "Referencia"]
    ws_ing.append(headers_ing)
    for cell in ws_ing[1]:
        cell.fill = green_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for p in payments:
        customer = p.sale.customer_name if p.sale else ''
        ws_ing.append([
            p.date_created.strftime('%Y-%m-%d %H:%M'),
            customer,
            float(p.amount),
            p.payment_via,
            p.payment_reference or ''
        ])
        ws_ing.cell(row=ws_ing.max_row, column=3).number_format = money_fmt

    # === Hoja 3: Egresos (detalle) ===
    ws_egr = wb.create_sheet("Egresos")
    headers_egr = ["Fecha Recepción", "Proveedor", "N° Factura", "Monto"]
    ws_egr.append(headers_egr)
    for cell in ws_egr[1]:
        cell.fill = red_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for o in orders:
        ws_egr.append([
            o.date_received.strftime('%Y-%m-%d %H:%M') if o.date_received else '',
            o.supplier_name or '',
            o.invoice_number or '',
            float(o.total) if o.total else 0
        ])
        ws_egr.cell(row=ws_egr.max_row, column=4).number_format = money_fmt

    # Auto-width en todas las hojas
    for ws in [ws_resumen, ws_ing, ws_egr]:
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass
            ws.column_dimensions[column].width = min(40, max_length + 2)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"flujo_caja_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
