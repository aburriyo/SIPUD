"""
Warehouse Operations Blueprint
Gestiona operaciones diarias del almacén: pedidos, recepciones, mermas
"""

from flask import Blueprint, render_template, request, jsonify, redirect, url_for, g, abort, send_file
from flask_login import login_required, current_user
from app.models import Product, InboundOrder, InboundOrderLineItem, Wastage, Lot, Supplier, ProductBundle, ActivityLog, utc_now
from datetime import datetime, timedelta
from bson import ObjectId
from mongoengine import DoesNotExist
import uuid
import io
from functools import wraps

bp = Blueprint("warehouse", __name__, url_prefix="/warehouse")


def permission_required(module, action='view'):
    """Decorator to check permissions before accessing a route"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                abort(401)
            if not current_user.has_permission(module, action):
                abort(403)
            return f(*args, **kwargs)
        return decorated_function
    return decorator


@bp.route("/")
@bp.route("/dashboard")
@login_required
def dashboard():
    """Dashboard de operaciones de almacén"""
    tenant = g.current_tenant

    # Productos próximos a vencer (30 días)
    thirty_days = datetime.now() + timedelta(days=30)
    expiring_soon = Product.objects(
        tenant=tenant,
        expiry_date__ne=None,
        expiry_date__lte=thirty_days.date(),
        expiry_date__gte=datetime.now().date()
    ).order_by('expiry_date').limit(10)

    # Productos con stock crítico
    all_products = Product.objects(tenant=tenant)

    # Filtrar y ordenar en Python ya que total_stock es una propiedad calculada
    low_stock = sorted(
        [p for p in all_products if p.total_stock <= p.critical_stock],
        key=lambda x: x.total_stock,
    )[:10]

    # Pedidos pendientes de recepción
    pending_orders = InboundOrder.objects(
        tenant=tenant,
        status__in=['pending', 'partially_received']
    ).order_by('-created_at').limit(10)

    return render_template(
        "warehouse/dashboard.html",
        expiring_soon=expiring_soon,
        low_stock=low_stock,
        pending_orders=pending_orders,
    )


@bp.route("/orders")
@login_required
def orders():
    """Gestión de pedidos a proveedores"""
    tenant = g.current_tenant
    orders = InboundOrder.objects(tenant=tenant).order_by('-created_at')
    return render_template("warehouse/orders.html", orders=orders)


@bp.route("/receiving")
@login_required
def receiving():
    """Recepción de mercancía"""
    tenant = g.current_tenant
    pending = InboundOrder.objects(
        tenant=tenant,
        status__in=['pending', 'partially_received']
    ).order_by('-created_at')
    return render_template("warehouse/receiving.html", pending_orders=pending)


@bp.route("/wastage")
@login_required
def wastage():
    """Registro de mermas"""
    tenant = g.current_tenant
    products = Product.objects(tenant=tenant).order_by('name')
    return render_template("warehouse/wastage.html", products=products)


@bp.route("/expiry")
@login_required
def expiry():
    """Gestión de vencimientos"""
    tenant = g.current_tenant
    products = Product.objects(tenant=tenant).order_by('expiry_date')
    return render_template(
        "warehouse/expiry.html", products=products, now=datetime.now().date()
    )


def generate_lot_code(supplier=None, product=None):
    """Genera código de lote legible: LOT-{PROV}-{SKU}-{YYMMDD}-{SUFFIX}"""
    now = datetime.now()
    date_part = now.strftime('%y%m%d')

    prov_part = 'GEN'
    if supplier:
        if hasattr(supplier, 'abbreviation') and supplier.abbreviation:
            prov_part = supplier.abbreviation.upper()[:4]
        elif hasattr(supplier, 'name') and supplier.name:
            prov_part = supplier.name.upper()[:4].replace(' ', '')
        elif isinstance(supplier, str) and supplier:
            prov_part = supplier.upper()[:4].replace(' ', '')

    sku_part = 'PROD'
    if product:
        if hasattr(product, 'sku') and product.sku:
            sku_part = product.sku.upper()[:8]
        elif hasattr(product, 'name') and product.name:
            sku_part = product.name.upper()[:6].replace(' ', '')

    suffix = uuid.uuid4().hex[:4].upper()
    return f"LOT-{prov_part}-{sku_part}-{date_part}-{suffix}"


# Supplier API
@bp.route("/api/suppliers", methods=["GET"])
@login_required
@permission_required('orders', 'view')
def get_suppliers():
    """Listar proveedores activos con búsqueda opcional"""
    try:
        tenant = g.current_tenant
        q = request.args.get('q', '').strip()

        query = Supplier.objects(tenant=tenant, is_active__ne=False)
        if q:
            query = query.filter(name__icontains=q)

        suppliers = query.order_by('name')

        return jsonify({
            "success": True,
            "suppliers": [{
                "id": str(s.id),
                "name": s.name,
                "rut": s.rut or '',
                "contact_info": s.contact_info or '',
                "abbreviation": s.abbreviation or ''
            } for s in suppliers]
        }), 200
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/suppliers", methods=["POST"])
@login_required
@permission_required('orders', 'create')
def create_supplier():
    """Crear nuevo proveedor"""
    try:
        data = request.get_json()
        tenant = g.current_tenant

        if not data:
            return jsonify({"success": False, "error": "No se recibieron datos"}), 400

        name = data.get("name", "").strip()
        if not name:
            return jsonify({"success": False, "error": "El nombre es obligatorio"}), 400

        rut = data.get("rut", "").strip() or None

        supplier = Supplier(
            name=name,
            rut=rut,
            contact_info=data.get("contact_info", "").strip(),
            abbreviation=data.get("abbreviation", "").strip()[:10] or None,
            tenant=tenant
        )
        supplier.save()

        ActivityLog.log(
            user=current_user,
            action='create',
            module='orders',
            description=f'Creó proveedor "{name}"',
            target_id=str(supplier.id),
            target_type='Supplier',
            request=request,
            tenant=tenant
        )

        return jsonify({
            "success": True,
            "message": "Proveedor creado exitosamente",
            "supplier": {
                "id": str(supplier.id),
                "name": supplier.name,
                "rut": supplier.rut or '',
                "contact_info": supplier.contact_info or '',
                "abbreviation": supplier.abbreviation or ''
            }
        }), 201
    except Exception as e:
        error_msg = str(e)
        if 'duplicate' in error_msg.lower() or 'unique' in error_msg.lower():
            return jsonify({"success": False, "error": "Ya existe un proveedor con ese RUT"}), 400
        return jsonify({"success": False, "error": error_msg}), 500


@bp.route("/api/suppliers/template", methods=["GET"])
@login_required
@permission_required('orders', 'view')
def download_supplier_template():
    """Descargar plantilla Excel para registro masivo de proveedores"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Proveedores"

    headers = ["Nombre (*)", "RUT", "Contacto", "Abreviación"]
    ws.append(headers)

    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        bottom=Side(style='thin', color='94A3B8')
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Ejemplo para guiar al usuario
    ws.append(["Distribuidora Ejemplo SpA", "76.123.456-7", "contacto@ejemplo.cl / +56912345678", "DEJE"])
    for cell in ws[2]:
        cell.font = Font(italic=True, color="94A3B8")
        cell.border = thin_border

    # Instrucciones en fila 4
    ws.append([])
    ws.append(["INSTRUCCIONES:"])
    ws['A4'].font = Font(bold=True, size=10, color="DC2626")
    ws.append(["- Nombre (*) es obligatorio. Los demás campos son opcionales."])
    ws.append(["- RUT debe ser único por proveedor (ej: 76.123.456-7)."])
    ws.append(["- Abreviación máx 10 caracteres, se usa para códigos de lote (ej: COSM, DIST)."])
    ws.append(["- Elimina esta fila de ejemplo antes de subir."])

    for row in range(5, 9):
        ws.cell(row=row, column=1).font = Font(size=9, color="64748B")

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 16

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="plantilla_proveedores.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.route("/api/suppliers/upload", methods=["POST"])
@login_required
@permission_required('orders', 'create')
def upload_suppliers():
    """Subir Excel con proveedores para registro masivo"""
    from openpyxl import load_workbook

    try:
        tenant = g.current_tenant

        if 'file' not in request.files:
            return jsonify({"success": False, "error": "No se recibió archivo"}), 400

        file = request.files['file']
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({"success": False, "error": "El archivo debe ser Excel (.xlsx)"}), 400

        wb = load_workbook(file, read_only=True)
        ws = wb.active

        created = []
        errors = []
        row_num = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            row_num += 1

            # Saltar filas vacías o de instrucciones
            if not row or not row[0]:
                continue
            name = str(row[0]).strip()
            if not name or name.startswith("INSTRUCCIONES") or name.startswith("-"):
                continue

            rut = str(row[1]).strip() if len(row) > 1 and row[1] else None
            contact_info = str(row[2]).strip() if len(row) > 2 and row[2] else ''
            abbreviation = str(row[3]).strip()[:10] if len(row) > 3 and row[3] else None

            # Verificar duplicado por nombre en este tenant
            existing = Supplier.objects(tenant=tenant, name__iexact=name).first()
            if existing:
                errors.append(f"Fila {row_num + 1}: '{name}' ya existe, se omitió")
                continue

            # Verificar duplicado por RUT si se proporcionó
            if rut:
                existing_rut = Supplier.objects(rut=rut).first()
                if existing_rut:
                    errors.append(f"Fila {row_num + 1}: RUT '{rut}' ya registrado, se omitió")
                    continue

            try:
                supplier = Supplier(
                    name=name,
                    rut=rut,
                    contact_info=contact_info,
                    abbreviation=abbreviation,
                    tenant=tenant
                )
                supplier.save()
                created.append(name)
            except Exception as e:
                errors.append(f"Fila {row_num + 1}: Error con '{name}' - {str(e)}")

        wb.close()

        if created:
            ActivityLog.log(
                user=current_user,
                action='create',
                module='orders',
                description=f'Carga masiva: {len(created)} proveedores creados',
                target_type='Supplier',
                request=request,
                tenant=tenant
            )

        return jsonify({
            "success": True,
            "message": f"{len(created)} proveedores creados exitosamente",
            "created": created,
            "errors": errors
        }), 200

    except Exception as e:
        return jsonify({"success": False, "error": f"Error procesando archivo: {str(e)}"}), 500


# API Endpoints
@bp.route("/api/orders", methods=["GET"])
@login_required
@permission_required('orders', 'view')
def get_orders():
    """Obtener todos los pedidos"""
    try:
        tenant = g.current_tenant
        orders = InboundOrder.objects(tenant=tenant).order_by('-created_at')

        def serialize_order(o):
            data = {
                "id": str(o.id),
                "supplier": o.supplier_name,
                "supplier_id": str(o.supplier.id) if o.supplier else None,
                "invoice_number": o.invoice_number,
                "status": o.status,
                "total": float(o.total) if o.total else 0,
                "notes": o.notes,
                "date_received": o.date_received.strftime("%d/%m/%Y %H:%M") if o.date_received else "",
                "created_at": o.created_at.strftime("%d/%m/%Y %H:%M") if o.created_at else "",
                "line_items": []
            }
            if o.line_items:
                for li in o.line_items:
                    data["line_items"].append({
                        "product_id": str(li.product.id) if li.product else None,
                        "product_name": li.product_name,
                        "product_sku": li.product_sku or '',
                        "quantity_ordered": li.quantity_ordered,
                        "quantity_received": li.quantity_received or 0,
                        "unit_cost": float(li.unit_cost) if li.unit_cost else 0
                    })
            return data

        return jsonify({
            "success": True,
            "orders": [serialize_order(o) for o in orders]
        }), 200

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/orders", methods=["POST"])
@login_required
@permission_required('orders', 'create')
def create_order():
    """Crear nuevo pedido a proveedor con line items opcionales"""
    try:
        data = request.get_json()
        tenant = g.current_tenant

        if not data:
            return jsonify({"success": False, "error": "No se recibieron datos"}), 400

        supplier_name = data.get("supplier", "").strip()
        invoice_number = data.get("invoice_number", "").strip()
        total = data.get("total", 0)

        if not supplier_name:
            return jsonify({"success": False, "error": "El proveedor es obligatorio"}), 400

        if not invoice_number:
            return jsonify({"success": False, "error": "El número de factura es obligatorio"}), 400

        try:
            total = float(total)
            if total < 0:
                return jsonify({"success": False, "error": "El total no puede ser negativo"}), 400
        except (ValueError, TypeError):
            return jsonify({"success": False, "error": "El total debe ser un número válido"}), 400

        # Resolver proveedor desde supplier_id si viene
        supplier_ref = None
        supplier_id = data.get("supplier_id")
        if supplier_id:
            try:
                supplier_ref = Supplier.objects.get(id=ObjectId(supplier_id), tenant=tenant)
                supplier_name = supplier_ref.name
            except DoesNotExist:
                pass

        # Procesar line items
        line_items = []
        items_data = data.get("items", [])
        for idx, item in enumerate(items_data):
            product_id = item.get("product_id")
            if not product_id:
                return jsonify({"success": False, "error": f"Item {idx + 1}: producto requerido"}), 400

            try:
                product = Product.objects.get(id=ObjectId(product_id), tenant=tenant)
            except DoesNotExist:
                return jsonify({"success": False, "error": f"Item {idx + 1}: producto no encontrado"}), 404

            qty = int(item.get("quantity_ordered", 0))
            if qty <= 0:
                return jsonify({"success": False, "error": f"Item {idx + 1}: cantidad debe ser mayor a 0"}), 400

            cost = float(item.get("unit_cost", 0))
            if cost < 0:
                return jsonify({"success": False, "error": f"Item {idx + 1}: costo no puede ser negativo"}), 400

            line_items.append(InboundOrderLineItem(
                product=product,
                product_name=product.name,
                product_sku=product.sku or '',
                quantity_ordered=qty,
                quantity_received=0,
                unit_cost=cost
            ))

        # Auto-calcular total desde line items si los hay y total es 0
        if line_items and total == 0:
            total = sum(li.quantity_ordered * float(li.unit_cost or 0) for li in line_items)

        new_order = InboundOrder(
            supplier=supplier_ref,
            supplier_name=supplier_name,
            invoice_number=invoice_number,
            notes=data.get("notes", "").strip(),
            status="pending",
            total=total,
            line_items=line_items,
            created_at=datetime.now(),
            tenant=tenant,
        )
        new_order.save()

        ActivityLog.log(
            user=current_user,
            action='create',
            module='orders',
            description=f'Creó pedido a "{supplier_name}" - Factura: {invoice_number}, Total: ${total:,.0f}' +
                        (f', {len(line_items)} producto(s)' if line_items else ''),
            target_id=str(new_order.id),
            target_type='InboundOrder',
            request=request,
            tenant=tenant
        )

        return jsonify({
            "success": True,
            "message": "Pedido creado exitosamente",
            "order_id": str(new_order.id),
        }), 201

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/orders/<order_id>", methods=["PUT"])
@login_required
@permission_required('orders', 'edit')
def update_order(order_id):
    """Actualizar pedido existente"""
    try:
        tenant = g.current_tenant
        try:
            order = InboundOrder.objects.get(id=ObjectId(order_id))
        except DoesNotExist:
            return jsonify({"success": False, "error": "Pedido no encontrado"}), 404

        if order.tenant != tenant:
            return jsonify({"success": False, "error": "Acceso denegado"}), 403

        data = request.get_json()

        if "supplier" in data:
            order.supplier_name = data["supplier"]
        if "supplier_id" in data and data["supplier_id"]:
            try:
                supplier_ref = Supplier.objects.get(id=ObjectId(data["supplier_id"]), tenant=tenant)
                order.supplier = supplier_ref
                order.supplier_name = supplier_ref.name
            except DoesNotExist:
                pass
        if "invoice_number" in data:
            order.invoice_number = data["invoice_number"]
        if "notes" in data:
            order.notes = data["notes"]
        if "total" in data:
            order.total = data["total"]
        if "status" in data:
            order.status = data["status"]

        # Editar line items solo si orden está pendiente
        if "items" in data and order.status == 'pending':
            line_items = []
            for idx, item in enumerate(data["items"]):
                product_id = item.get("product_id")
                if not product_id:
                    continue
                try:
                    product = Product.objects.get(id=ObjectId(product_id), tenant=tenant)
                except DoesNotExist:
                    return jsonify({"success": False, "error": f"Item {idx + 1}: producto no encontrado"}), 404

                qty = int(item.get("quantity_ordered", 0))
                if qty <= 0:
                    return jsonify({"success": False, "error": f"Item {idx + 1}: cantidad debe ser mayor a 0"}), 400

                cost = float(item.get("unit_cost", 0))
                line_items.append(InboundOrderLineItem(
                    product=product,
                    product_name=product.name,
                    product_sku=product.sku or '',
                    quantity_ordered=qty,
                    quantity_received=0,
                    unit_cost=cost
                ))
            order.line_items = line_items
            if line_items and float(data.get("total", 0)) == 0:
                order.total = sum(li.quantity_ordered * float(li.unit_cost or 0) for li in line_items)

        order.save()

        # Log activity
        ActivityLog.log(
            user=current_user,
            action='update',
            module='orders',
            description=f'Actualizó pedido de "{order.supplier_name}" - Factura: {order.invoice_number}',
            target_id=str(order.id),
            target_type='InboundOrder',
            request=request,
            tenant=tenant
        )

        return jsonify(
            {"success": True, "message": "Pedido actualizado exitosamente"}
        ), 200

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/orders/<order_id>", methods=["DELETE"])
@login_required
@permission_required('orders', 'delete')
def delete_order(order_id):
    """Eliminar pedido"""
    try:
        tenant = g.current_tenant
        try:
            order = InboundOrder.objects.get(id=ObjectId(order_id))
        except DoesNotExist:
            return jsonify({"success": False, "error": "Pedido no encontrado"}), 404

        if order.tenant != tenant:
            return jsonify({"success": False, "error": "Acceso denegado"}), 403

        # Verificar que no tenga lotes asociados
        if Lot.objects(order=order).count() > 0:
            return jsonify(
                {
                    "success": False,
                    "error": "No se puede eliminar un pedido con lotes asociados",
                }
            ), 400

        # Log activity before deletion
        supplier_name = order.supplier_name
        invoice_number = order.invoice_number
        order_id_str = str(order.id)

        order.delete()

        ActivityLog.log(
            user=current_user,
            action='delete',
            module='orders',
            description=f'Eliminó pedido de "{supplier_name}" - Factura: {invoice_number}',
            target_id=order_id_str,
            target_type='InboundOrder',
            request=request,
            tenant=tenant
        )

        return jsonify(
            {"success": True, "message": "Pedido eliminado exitosamente"}
        ), 200

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/receiving/orders", methods=["GET"])
@login_required
def get_receiving_orders():
    """Obtener pedidos pendientes de recepción (pending y partially_received)"""
    try:
        tenant = g.current_tenant
        orders = InboundOrder.objects(
            tenant=tenant,
            status__in=['pending', 'partially_received']
        ).order_by('-created_at')

        result = []
        for o in orders:
            order_data = {
                "id": str(o.id),
                "supplier": o.supplier_name,
                "invoice_number": o.invoice_number,
                "status": o.status,
                "total": float(o.total) if o.total else 0,
                "notes": o.notes,
                "created_at": o.created_at.strftime("%d/%m/%Y %H:%M") if o.created_at else "",
                "line_items": []
            }
            if o.line_items:
                for li in o.line_items:
                    order_data["line_items"].append({
                        "product_id": str(li.product.id) if li.product else None,
                        "product_name": li.product_name,
                        "product_sku": li.product_sku or '',
                        "quantity_ordered": li.quantity_ordered,
                        "quantity_received": li.quantity_received or 0,
                        "unit_cost": float(li.unit_cost) if li.unit_cost else 0
                    })
            result.append(order_data)

        return jsonify({"success": True, "orders": result}), 200

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/receiving/<order_id>", methods=["POST"])
@login_required
@permission_required('orders', 'receive')
def confirm_receiving(order_id):
    """Confirmar recepción de pedido con soporte para recepción parcial"""
    try:
        tenant = g.current_tenant
        try:
            order = InboundOrder.objects.get(id=ObjectId(order_id))
        except DoesNotExist:
            return jsonify({"success": False, "error": "Pedido no encontrado"}), 404

        if order.tenant != tenant:
            return jsonify({"success": False, "error": "Acceso denegado"}), 403

        if order.status in ("received", "paid"):
            return jsonify({"success": False, "error": "Este pedido ya fue recibido completamente"}), 400

        data = request.get_json()
        if not data:
            return jsonify({"success": False, "error": "No se recibieron datos"}), 400

        products = data.get("products", [])
        if not products:
            return jsonify({"success": False, "error": "Debe agregar al menos un producto a la recepción"}), 400

        # Resolver supplier para lot codes
        supplier_for_code = order.supplier or order.supplier_name

        created_lots = []

        for idx, item in enumerate(products):
            if not item.get("product_id"):
                return jsonify({"success": False, "error": f"Producto {idx + 1}: Debe seleccionar un producto"}), 400

            try:
                quantity = int(item.get("quantity", 0))
                if quantity <= 0:
                    return jsonify({"success": False, "error": f"Producto {idx + 1}: La cantidad debe ser mayor a 0"}), 400
            except (ValueError, TypeError):
                return jsonify({"success": False, "error": f"Producto {idx + 1}: Cantidad inválida"}), 400

            try:
                product = Product.objects.get(id=ObjectId(item["product_id"]))
            except DoesNotExist:
                return jsonify({"success": False, "error": f"Producto {idx + 1}: Producto no encontrado"}), 404

            if product.tenant != tenant:
                return jsonify({"success": False, "error": f"Producto {idx + 1}: Acceso denegado"}), 403

            # Generar lot code legible si no viene
            lot_code = str(item.get("lot_code", "")).strip()
            if not lot_code:
                lot_code = generate_lot_code(supplier=supplier_for_code, product=product)

            # Unit cost
            unit_cost = 0
            if item.get("unit_cost") is not None:
                try:
                    unit_cost = float(item["unit_cost"])
                except (ValueError, TypeError):
                    unit_cost = 0

            # Si la orden tiene line_items, buscar el line item correspondiente para heredar costo
            if not unit_cost and order.line_items:
                for li in order.line_items:
                    if li.product and str(li.product.id) == item["product_id"]:
                        unit_cost = float(li.unit_cost or 0)
                        break

            # Fecha de vencimiento
            expiry_date = None
            if item.get("expiry_date"):
                try:
                    expiry_date = datetime.strptime(str(item["expiry_date"]).strip(), "%Y-%m-%d").date()
                    if expiry_date < datetime.now().date():
                        return jsonify({"success": False, "error": f"Producto {idx + 1}: La fecha de vencimiento no puede ser en el pasado"}), 400
                except (ValueError, TypeError):
                    return jsonify({"success": False, "error": f"Producto {idx + 1}: Formato de fecha inválido (use YYYY-MM-DD)"}), 400

            lot = Lot(
                product=product,
                order=order,
                tenant=tenant,
                lot_code=lot_code,
                quantity_initial=quantity,
                quantity_current=quantity,
                unit_cost=unit_cost,
                expiry_date=expiry_date,
            )
            lot.save()

            created_lots.append({
                "lot_id": str(lot.id),
                "lot_code": lot_code,
                "product_name": product.name,
                "product_sku": product.sku or '',
                "quantity": quantity,
                "unit_cost": unit_cost,
                "expiry_date": expiry_date.strftime("%d/%m/%Y") if expiry_date else None
            })

            # Actualizar quantity_received en line_items si existen
            if order.line_items:
                for li in order.line_items:
                    if li.product and str(li.product.id) == item["product_id"]:
                        li.quantity_received = (li.quantity_received or 0) + quantity
                        break

        # Determinar status final
        if order.line_items:
            all_received = all(
                (li.quantity_received or 0) >= li.quantity_ordered
                for li in order.line_items
            )
            order.status = "received" if all_received else "partially_received"
        else:
            order.status = "received"

        order.date_received = datetime.now()
        order.save()

        ActivityLog.log(
            user=current_user,
            action='receive',
            module='orders',
            description=f'Recibió {"parcialmente " if order.status == "partially_received" else ""}pedido de "{order.supplier_name}" - Factura: {order.invoice_number}, {len(products)} producto(s)',
            target_id=str(order.id),
            target_type='InboundOrder',
            request=request,
            tenant=tenant
        )

        status_label = "parcialmente recibido" if order.status == "partially_received" else "recibido"
        return jsonify({
            "success": True,
            "message": f"Recepción confirmada. Pedido {status_label}. {len(products)} producto(s) agregado(s) al inventario.",
            "status": order.status,
            "lots": created_lots
        }), 200

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/orders/<order_id>/receiving-summary", methods=["GET"])
@login_required
@permission_required('orders', 'view')
def receiving_summary(order_id):
    """Resumen de recepción: lotes creados, cantidades y costos"""
    try:
        tenant = g.current_tenant
        try:
            order = InboundOrder.objects.get(id=ObjectId(order_id))
        except DoesNotExist:
            return jsonify({"success": False, "error": "Pedido no encontrado"}), 404

        if order.tenant != tenant:
            return jsonify({"success": False, "error": "Acceso denegado"}), 403

        lots = Lot.objects(order=order)
        lots_data = [{
            "lot_code": lot.lot_code,
            "product_name": lot.product.name if lot.product else 'N/A',
            "product_sku": lot.product.sku if lot.product else '',
            "quantity": lot.quantity_initial,
            "unit_cost": float(lot.unit_cost) if lot.unit_cost else 0,
            "subtotal": lot.quantity_initial * float(lot.unit_cost or 0),
            "expiry_date": lot.expiry_date.strftime("%d/%m/%Y") if lot.expiry_date else None,
            "created_at": lot.created_at.strftime("%d/%m/%Y %H:%M") if lot.created_at else ''
        } for lot in lots]

        total_items = sum(l["quantity"] for l in lots_data)
        total_cost = sum(l["subtotal"] for l in lots_data)

        return jsonify({
            "success": True,
            "order_id": str(order.id),
            "supplier": order.supplier_name,
            "invoice_number": order.invoice_number,
            "status": order.status,
            "lots": lots_data,
            "total_items": total_items,
            "total_cost": total_cost
        }), 200

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/wastage", methods=["POST"])
@login_required
@permission_required('wastage', 'create')
def register_wastage():
    """Registrar merma de producto"""
    try:
        data = request.get_json()
        tenant = g.current_tenant

        # Validaciones
        if not data:
            return jsonify({"success": False, "error": "No se recibieron datos"}), 400

        product_id = data.get("product_id")
        quantity_str = data.get("quantity", 0)
        reason = data.get("reason", "").strip()

        if not product_id:
            return jsonify(
                {"success": False, "error": "Debe seleccionar un producto"}
            ), 400

        try:
            quantity = int(quantity_str)
            if quantity <= 0:
                return jsonify(
                    {"success": False, "error": "La cantidad debe ser mayor a 0"}
                ), 400
        except (ValueError, TypeError):
            return jsonify(
                {
                    "success": False,
                    "error": "La cantidad debe ser un número entero válido",
                }
            ), 400

        if not reason or reason not in ["vencido", "dañado", "perdido", "robo", "otro"]:
            return jsonify(
                {"success": False, "error": "Debe especificar una razón válida"}
            ), 400

        # Validar producto
        try:
            product = Product.objects.get(id=ObjectId(product_id))
        except DoesNotExist:
            return jsonify({"success": False, "error": "Producto no encontrado"}), 404

        if product.tenant != tenant:
            return jsonify({"success": False, "error": "Acceso denegado"}), 403

        # Verificar stock disponible
        if product.total_stock < quantity:
            return jsonify(
                {
                    "success": False,
                    "error": f"Stock insuficiente. Disponible: {product.total_stock}, solicitado: {quantity}",
                }
            ), 400

        # Registrar merma
        wastage_record = Wastage(
            product=product,
            quantity=quantity,
            reason=reason,
            notes=data.get("notes", "").strip(),
            tenant=tenant,
        )

        # Reducir stock (FIFO - del lote más antiguo)
        remaining_to_deduct = quantity
        available_lots = sorted(
            [l for l in product.lots if l.quantity_current > 0],
            key=lambda x: x.created_at,
        )

        if not available_lots:
            return jsonify(
                {
                    "success": False,
                    "error": "No hay lotes disponibles para este producto",
                }
            ), 400

        for lot in available_lots:
            if remaining_to_deduct <= 0:
                break
            deduct = min(lot.quantity_current, remaining_to_deduct)
            lot.quantity_current -= deduct
            remaining_to_deduct -= deduct
            lot.save()

        wastage_record.save()

        # Log activity
        reason_names = {
            'vencido': 'Vencido',
            'dañado': 'Dañado',
            'perdido': 'Perdido',
            'robo': 'Robo',
            'otro': 'Otro'
        }
        ActivityLog.log(
            user=current_user,
            action='create',
            module='wastage',
            description=f'Registró merma de "{product.name}" - {quantity} unidades ({reason_names.get(reason, reason)})',
            target_id=str(wastage_record.id),
            target_type='Wastage',
            request=request,
            tenant=tenant
        )

        return jsonify(
            {
                "success": True,
                "message": "Merma registrada exitosamente",
                "wastage_id": str(wastage_record.id),
            }
        ), 201

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/wastage/history", methods=["GET"])
@login_required
@permission_required('wastage', 'view')
def wastage_history():
    """Obtener historial de mermas"""
    try:
        tenant = g.current_tenant
        wastages = Wastage.objects(tenant=tenant).order_by('-date_created')

        return jsonify(
            {
                "success": True,
                "wastages": [
                    {
                        "id": str(w.id),
                        "product_id": str(w.product.id) if w.product else None,
                        "product_name": w.product.name if w.product else 'N/A',
                        "quantity": w.quantity,
                        "reason": w.reason,
                        "notes": w.notes,
                        "date": w.date_created.strftime("%d/%m/%Y %H:%M"),
                    }
                    for w in wastages
                ],
            }
        ), 200

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/wastage/<wastage_id>", methods=["DELETE"])
@login_required
@permission_required('wastage', 'delete')
def delete_wastage(wastage_id):
    """Eliminar registro de merma (NO REVIERTE STOCK)"""
    try:
        tenant = g.current_tenant
        try:
            wastage_record = Wastage.objects.get(id=ObjectId(wastage_id))
        except DoesNotExist:
            return jsonify({"success": False, "error": "Merma no encontrada"}), 404

        if wastage_record.tenant != tenant:
            return jsonify({"success": False, "error": "Acceso denegado"}), 403

        # Log activity before deletion
        product_name = wastage_record.product.name if wastage_record.product else 'N/A'
        quantity = wastage_record.quantity
        wastage_id_str = str(wastage_record.id)

        wastage_record.delete()

        ActivityLog.log(
            user=current_user,
            action='delete',
            module='wastage',
            description=f'Eliminó registro de merma de "{product_name}" - {quantity} unidades',
            target_id=wastage_id_str,
            target_type='Wastage',
            request=request,
            tenant=tenant
        )

        return jsonify({"success": True, "message": "Registro de merma eliminado"}), 200

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/expiry/products", methods=["GET"])
@login_required
def get_expiry_products():
    """Obtener productos con fechas de vencimiento"""
    try:
        tenant = g.current_tenant
        products = Product.objects(
            tenant=tenant,
            expiry_date__ne=None
        ).order_by('expiry_date')

        now = datetime.now().date()

        return jsonify(
            {
                "success": True,
                "products": [
                    {
                        "id": str(p.id),
                        "name": p.name,
                        "sku": p.sku,
                        "expiry_date": p.expiry_date.strftime("%Y-%m-%d")
                        if p.expiry_date
                        else None,
                        "days_until_expiry": (p.expiry_date - now).days
                        if p.expiry_date
                        else None,
                        "status": "expired"
                        if p.expiry_date and p.expiry_date < now
                        else (
                            "warning"
                            if p.expiry_date and (p.expiry_date - now).days <= 30
                            else "ok"
                        ),
                        "stock": p.total_stock,
                    }
                    for p in products
                ],
            }
        ), 200

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/expiry/<product_id>", methods=["PUT"])
@login_required
def update_expiry(product_id):
    """Actualizar fecha de vencimiento de producto"""
    try:
        data = request.get_json()
        tenant = g.current_tenant

        if not data:
            return jsonify({"success": False, "error": "No se recibieron datos"}), 400

        try:
            product = Product.objects.get(id=ObjectId(product_id))
        except DoesNotExist:
            return jsonify({"success": False, "error": "Producto no encontrado"}), 404

        if product.tenant != tenant:
            return jsonify({"success": False, "error": "Acceso denegado"}), 403

        # Validar y actualizar fecha de vencimiento
        expiry_str = data.get("expiry_date", "").strip()

        if expiry_str:
            try:
                expiry_date = datetime.strptime(expiry_str, "%Y-%m-%d").date()

                # Validar que la fecha no esté en el pasado
                if expiry_date < datetime.now().date():
                    return jsonify(
                        {
                            "success": False,
                            "error": "La fecha de vencimiento no puede estar en el pasado",
                        }
                    ), 400

                product.expiry_date = expiry_date
            except ValueError:
                return jsonify(
                    {
                        "success": False,
                        "error": "Formato de fecha inválido. Use YYYY-MM-DD",
                    }
                ), 400
        else:
            product.expiry_date = None

        product.save()

        # Log activity
        ActivityLog.log(
            user=current_user,
            action='update',
            module='products',
            description=f'Actualizó fecha de vencimiento de "{product.name}" a {expiry_str or "sin fecha"}',
            target_id=str(product.id),
            target_type='Product',
            request=request,
            tenant=tenant
        )

        return jsonify(
            {"success": True, "message": "Fecha de vencimiento actualizada"}
        ), 200

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/alerts", methods=["GET"])
@login_required
def get_alerts():
    """Obtener alertas de vencimientos próximos y stock crítico"""
    try:
        tenant = g.current_tenant
        alerts = []
        now = datetime.now().date()

        # Alertas de vencimiento (próximos 30 días)
        thirty_days = now + timedelta(days=30)
        expiring_products = Product.objects(
            tenant=tenant,
            expiry_date__ne=None,
            expiry_date__lte=thirty_days,
            expiry_date__gte=now
        ).order_by('expiry_date')

        for product in expiring_products:
            days_left = (product.expiry_date - now).days
            severity = (
                "danger"
                if days_left <= 7
                else ("warning" if days_left <= 14 else "info")
            )

            alerts.append(
                {
                    "type": "expiry",
                    "severity": severity,
                    "product_id": str(product.id),
                    "product_name": product.name,
                    "message": f"El producto '{product.name}' vence en {days_left} días",
                    "days_left": days_left,
                    "expiry_date": product.expiry_date.strftime("%d/%m/%Y"),
                    "stock": product.total_stock,
                }
            )

        # Alertas de stock crítico
        all_products = Product.objects(tenant=tenant)
        for product in all_products:
            if product.total_stock <= product.critical_stock:
                alerts.append(
                    {
                        "type": "low_stock",
                        "severity": "danger" if product.total_stock == 0 else "warning",
                        "product_id": str(product.id),
                        "product_name": product.name,
                        "message": f"Stock bajo: '{product.name}' ({product.total_stock} unidades)",
                        "current_stock": product.total_stock,
                        "critical_stock": product.critical_stock,
                    }
                )

        # Ordenar por severidad (danger primero)
        severity_order = {"danger": 0, "warning": 1, "info": 2}
        alerts.sort(key=lambda x: severity_order.get(x["severity"], 3))

        return jsonify({"success": True, "alerts": alerts, "count": len(alerts)}), 200

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/assembly", methods=["POST"])
@login_required
@permission_required('orders', 'create')
def assemble_box():
    """
    Ensamblar cajas (Kitting).
    Descuenta stock de componentes y agrega stock al producto 'caja' (bundle).
    """
    try:
        data = request.get_json()
        tenant = g.current_tenant
        bundle_id = data.get("bundle_id")
        quantity = int(data.get("quantity", 0))

        if quantity <= 0:
            return jsonify(
                {"success": False, "error": "Cantidad debe ser mayor a 0"}
            ), 400

        # 1. Verificar producto bundle
        try:
            bundle_product = Product.objects.get(id=ObjectId(bundle_id))
        except DoesNotExist:
            return jsonify({"success": False, "error": "Producto no encontrado"}), 404

        if bundle_product.tenant != tenant:
            return jsonify({"success": False, "error": "Acceso denegado"}), 403

        if not bundle_product.is_bundle:
            return jsonify(
                {
                    "success": False,
                    "error": "El producto seleccionado no es un Kit/Caja",
                }
            ), 400

        # 2. Verificar componentes y stock
        components = ProductBundle.objects(bundle=bundle_product)

        if components.count() == 0:
            return jsonify(
                {"success": False, "error": "Este kit no tiene componentes definidos"}
            ), 400

        # Pre-chequeo de stock
        for comp_rel in components:
            comp_product = comp_rel.component
            needed_qty = comp_rel.quantity * quantity

            if comp_product.total_stock < needed_qty:
                return jsonify(
                    {
                        "success": False,
                        "error": f'Stock insuficiente de componente "{comp_product.name}". Requerido: {needed_qty}, Disponible: {comp_product.total_stock}',
                    }
                ), 400

        # 3. Descontar stock de componentes (FIFO)
        for comp_rel in components:
            comp_product = comp_rel.component
            total_needed = comp_rel.quantity * quantity
            remaining_to_deduct = total_needed

            # Obtener lotes disponibles ordenados por fecha (FIFO)
            available_lots = sorted(
                [l for l in comp_product.lots if l.quantity_current > 0],
                key=lambda x: x.created_at,
            )

            for lot in available_lots:
                if remaining_to_deduct <= 0:
                    break

                deduct = min(lot.quantity_current, remaining_to_deduct)
                lot.quantity_current -= deduct
                remaining_to_deduct -= deduct
                lot.save()

            if remaining_to_deduct > 0:
                raise Exception(
                    f"Error concurrente de inventario en {comp_product.name}"
                )

        # 4. Crear Lote para el Bundle
        # Buscamos o creamos una Orden de Entrada "interna" para asignar el lote
        internal_supplier_name = "Interno: Armado"
        today_assembly_order = InboundOrder.objects(
            tenant=tenant,
            supplier_name=internal_supplier_name,
            created_at__gte=datetime.combine(utc_now().date(), datetime.min.time()),
            created_at__lt=datetime.combine(utc_now().date() + timedelta(days=1), datetime.min.time())
        ).first()

        if not today_assembly_order:
            # Buscar o crear proveedor interno
            internal_supplier = Supplier.objects(
                rut="99999999-9", tenant=tenant
            ).first()
            if not internal_supplier:
                internal_supplier = Supplier.objects(
                    name="Interno", tenant=tenant
                ).first()

            if not internal_supplier:
                internal_supplier = Supplier(
                    name="Interno",
                    rut="99999999-9",
                    contact_info="Sistema",
                    tenant=tenant,
                )
                internal_supplier.save()

            today_assembly_order = InboundOrder(
                supplier=internal_supplier,
                supplier_name=internal_supplier_name,
                invoice_number=f"ARM-{utc_now().strftime('%Y%m%d-%H%M')}",
                status="received",
                date_received=utc_now(),
                notes="Generado automáticamente por proceso de armado de cajas",
                tenant=tenant,
            )
            today_assembly_order.save()

        # FIXED: Using 'order' instead of 'inbound_order_id' and adding tenant
        new_lot = Lot(
            product=bundle_product,
            order=today_assembly_order,
            tenant=tenant,  # FIXED: Added tenant
            lot_code=f"KIT-{utc_now().strftime('%Y%m%d%H%M%S')}",
            quantity_initial=quantity,
            quantity_current=quantity,
            expiry_date=None,
            created_at=utc_now(),
        )
        new_lot.save()

        # Log activity
        ActivityLog.log(
            user=current_user,
            action='create',
            module='orders',
            description=f'Armó {quantity} unidades de "{bundle_product.name}" (Kit/Bundle)',
            target_id=str(new_lot.id),
            target_type='Lot',
            request=request,
            tenant=tenant
        )

        return jsonify(
            {
                "success": True,
                "message": f'Se armaron {quantity} unidades de "{bundle_product.name}" exitosamente',
            }
        ), 201

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
