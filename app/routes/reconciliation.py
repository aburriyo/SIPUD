"""
Módulo de Cuadratura Bancaria - SIPUD
Permite importar cartolas bancarias y conciliar con ventas.
"""
import os
import logging
from flask import Blueprint, jsonify, request, g, render_template, send_file
from flask_login import login_required, current_user
from app.models import BankTransaction, Sale, Payment, Tenant, ActivityLog, utc_now
from datetime import datetime, timedelta
from decimal import Decimal
from bson import ObjectId
from functools import wraps
from io import BytesIO

logger = logging.getLogger(__name__)

bp = Blueprint('reconciliation', __name__, url_prefix='/reconciliation')


def permission_required(module, action='view'):
    """Decorator to check permissions using ROLE_PERMISSIONS"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                return jsonify({'error': 'No autenticado'}), 401
            if not current_user.has_permission(module, action):
                return jsonify({'error': 'No tienes permisos para esta acción'}), 403
            return f(*args, **kwargs)
        return decorated_function
    return decorator


@bp.route('/')
@login_required
def reconciliation_view():
    """Render reconciliation view page"""
    if not current_user.has_permission('reconciliation', 'view'):
        return "No tienes permisos para ver esta página", 403
    return render_template('reconciliation.html')


@bp.route('/api/transactions', methods=['GET'])
@login_required
@permission_required('reconciliation', 'view')
def get_transactions():
    """Get bank transactions with filters"""
    tenant = g.current_tenant
    
    # Query params
    status = request.args.get('status', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    q = request.args.get('q', '').strip()
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 50))

    # Build query
    query = BankTransaction.objects(tenant=tenant)

    if status:
        query = query.filter(status=status)

    if q:
        from mongoengine.queryset.visitor import Q
        query = query.filter(
            Q(description__icontains=q) | Q(reference__icontains=q)
        )

    if date_from:
        try:
            dt_from = datetime.strptime(date_from, '%Y-%m-%d')
            query = query.filter(date__gte=dt_from)
        except ValueError:
            pass

    if date_to:
        try:
            dt_to = datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(date__lt=dt_to)
        except ValueError:
            pass
    
    # Pagination
    total = query.count()
    transactions = query.order_by('-date').skip((page - 1) * per_page).limit(per_page)
    
    # Format results
    results = []
    for t in transactions:
        results.append({
            'id': str(t.id),
            'date': t.date.strftime('%Y-%m-%d') if t.date else None,
            'amount': float(t.amount) if t.amount else 0,
            'description': t.description or '',
            'reference': t.reference or '',
            'transaction_type': t.transaction_type,
            'status': t.status,
            'matched_sale_id': str(t.matched_sale.id) if t.matched_sale else None,
            'matched_sale_customer': t.matched_sale.customer_name if t.matched_sale else None,
            'match_type': t.match_type,
        })
    
    return jsonify({
        'transactions': results,
        'total': total,
        'page': page,
        'per_page': per_page,
        'pages': (total + per_page - 1) // per_page
    })


def _process_csv_file(file, tenant, user, req):
    """Process CSV bank statement file"""
    import csv
    from io import StringIO
    
    # Column detection aliases
    date_aliases = ['fecha', 'date', 'fec', 'fecha operación', 'fecha_operacion']
    amount_aliases = ['monto', 'amount', 'valor', 'importe', 'cargo', 'abono', 'total']
    desc_aliases = ['descripción', 'descripcion', 'description', 'detalle', 'glosa', 'concepto', 'movimiento']
    ref_aliases = ['referencia', 'reference', 'ref', 'número', 'numero', 'nro', 'operación', 'operacion', 'documento', 'n° documento']
    
    try:
        # Read file content and decode
        content = file.read()
        # Try different encodings
        for encoding in ['utf-8', 'latin-1', 'cp1252']:
            try:
                text = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            return jsonify({'error': 'No se pudo decodificar el archivo. Intenta guardarlo como UTF-8.'}), 400
        
        # Detect delimiter
        sample = text[:2000]
        if sample.count(';') > sample.count(','):
            delimiter = ';'
        else:
            delimiter = ','
        
        reader = csv.reader(StringIO(text), delimiter=delimiter)
        rows = list(reader)
        
        if len(rows) < 2:
            return jsonify({'error': 'El archivo está vacío o no tiene datos'}), 400
        
        # Find header row (search first 30 rows)
        header_row_idx = None
        headers = []
        col_map = {}
        
        for idx, row in enumerate(rows[:30]):
            row_values = [str(cell).lower().strip() for cell in row]
            
            temp_map = {}
            for col_idx, h in enumerate(row_values):
                if any(alias in h for alias in date_aliases) and 'date' not in temp_map:
                    temp_map['date'] = col_idx
                elif any(alias in h for alias in amount_aliases) and 'amount' not in temp_map:
                    temp_map['amount'] = col_idx
                elif any(alias in h for alias in desc_aliases) and 'description' not in temp_map:
                    temp_map['description'] = col_idx
                elif any(alias in h for alias in ref_aliases) and 'reference' not in temp_map:
                    temp_map['reference'] = col_idx
            
            if 'date' in temp_map and 'amount' in temp_map:
                header_row_idx = idx
                headers = row_values
                col_map = temp_map
                break
        
        if header_row_idx is None:
            return jsonify({
                'error': 'No se pudieron detectar las columnas de Fecha y Monto.',
                'hint': 'Asegúrate de que el CSV tenga encabezados con "Fecha" y "Monto"'
            }), 400
        
        # Process data rows
        created = 0
        duplicates = 0
        errors = []
        filename = file.filename

        for row_idx, row in enumerate(rows[header_row_idx + 1:], start=header_row_idx + 2):
            if not row or all(not cell.strip() for cell in row):
                continue

            try:
                # Parse date
                date_val = row[col_map['date']].strip() if col_map['date'] < len(row) else ''
                if not date_val:
                    continue

                tx_date = None
                for fmt in ['%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%Y/%m/%d', '%d.%m.%Y']:
                    try:
                        tx_date = datetime.strptime(date_val, fmt)
                        break
                    except Exception:
                        continue

                if not tx_date:
                    errors.append(f'Fila {row_idx}: formato de fecha no reconocido "{date_val}"')
                    continue

                # Parse amount
                amount_val = row[col_map['amount']].strip() if col_map['amount'] < len(row) else ''
                if not amount_val:
                    continue

                # Clean amount string
                clean = amount_val.replace('$', '').replace('.', '').replace(',', '.').replace(' ', '').strip()
                try:
                    amount = Decimal(clean)
                except Exception:
                    errors.append(f'Fila {row_idx}: monto inválido "{amount_val}"')
                    continue

                # Determine transaction type
                tx_type = 'credit' if amount > 0 else 'debit'
                amount = abs(amount)

                # Get description and reference
                description = row[col_map['description']].strip() if col_map.get('description') is not None and col_map['description'] < len(row) else ''
                reference = row[col_map['reference']].strip() if col_map.get('reference') is not None and col_map['reference'] < len(row) else ''

                # Check for duplicates
                existing = BankTransaction.objects(
                    tenant=tenant, date=tx_date, amount=amount, description=description[:500]
                ).first()
                if existing:
                    duplicates += 1
                    continue

                # Create transaction
                tx = BankTransaction(
                    date=tx_date,
                    amount=amount,
                    description=description[:500],
                    reference=reference[:100],
                    transaction_type=tx_type,
                    status='pending',
                    source_file=filename,
                    row_number=row_idx,
                    tenant=tenant
                )
                tx.save()
                created += 1

            except Exception as e:
                errors.append(f'Fila {row_idx}: {str(e)}')

        # Log activity
        ActivityLog.log(
            user=user,
            action='create',
            module='reconciliation',
            description=f'Importó {created} transacciones bancarias desde "{filename}"',
            details={'file': filename, 'created': created, 'duplicates': duplicates, 'errors': len(errors)},
            request=req,
            tenant=tenant
        )

        return jsonify({
            'success': True,
            'created': created,
            'duplicates': duplicates,
            'errors': errors[:20],
            'total_errors': len(errors)
        })
        
    except Exception as e:
        return jsonify({'error': f'Error al procesar CSV: {str(e)}'}), 500


@bp.route('/api/transactions/upload', methods=['POST'])
@login_required
@permission_required('reconciliation', 'create')
def upload_transactions():
    """Upload Excel file with bank transactions"""
    tenant = g.current_tenant
    
    if 'file' not in request.files:
        return jsonify({'error': 'No se proporcionó archivo'}), 400
    
    file = request.files['file']
    filename = file.filename.lower()
    
    if not filename.endswith(('.xlsx', '.xls', '.csv')):
        return jsonify({'error': 'El archivo debe ser Excel (.xlsx) o CSV (.csv)'}), 400
    
    try:
        # Handle CSV files
        if filename.endswith('.csv'):
            return _process_csv_file(file, g.current_tenant, current_user, request)
        
        # Handle Excel files
        from openpyxl import load_workbook
        
        wb = load_workbook(file, read_only=True)
        ws = wb.active
        
        # Column detection aliases
        date_aliases = ['fecha', 'date', 'fec', 'fecha operación', 'fecha_operacion']
        amount_aliases = ['monto', 'amount', 'valor', 'importe', 'cargo', 'abono', 'total']
        desc_aliases = ['descripción', 'descripcion', 'description', 'detalle', 'glosa', 'concepto', 'movimiento']
        ref_aliases = ['referencia', 'reference', 'ref', 'número', 'numero', 'nro', 'operación', 'operacion', 'documento', 'n° documento']
        
        # Search for header row (may not be row 1 - bank statements often have metadata first)
        header_row = None
        headers = []
        col_map = {}
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
            row_values = [str(cell).lower().strip() if cell else '' for cell in row]
            
            # Check if this row contains header keywords
            temp_map = {}
            for idx, h in enumerate(row_values):
                if any(alias in h for alias in date_aliases) and 'date' not in temp_map:
                    temp_map['date'] = idx
                elif any(alias in h for alias in amount_aliases) and 'amount' not in temp_map:
                    temp_map['amount'] = idx
                elif any(alias in h for alias in desc_aliases) and 'description' not in temp_map:
                    temp_map['description'] = idx
                elif any(alias in h for alias in ref_aliases) and 'reference' not in temp_map:
                    temp_map['reference'] = idx
            
            # If we found both date and amount columns, this is our header row
            if 'date' in temp_map and 'amount' in temp_map:
                header_row = row_idx
                headers = row_values
                col_map = temp_map
                break
        
        if header_row is None:
            return jsonify({
                'error': 'No se pudieron detectar las columnas de Fecha y Monto. Asegúrate de que el archivo tenga encabezados con "Fecha" y "Monto".',
                'hint': 'Se buscó en las primeras 30 filas'
            }), 400
        
        # Process rows starting after header
        created = 0
        duplicates = 0
        errors = []
        filename = file.filename
        data_start_row = header_row + 1

        for row_idx, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
            try:
                # Parse date
                date_val = row[col_map['date']]
                if not date_val:
                    continue

                if isinstance(date_val, datetime):
                    tx_date = date_val
                elif isinstance(date_val, str):
                    for fmt in ['%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%Y/%m/%d']:
                        try:
                            tx_date = datetime.strptime(date_val.strip(), fmt)
                            break
                        except Exception:
                            continue
                    else:
                        errors.append(f'Fila {row_idx}: formato de fecha no reconocido')
                        continue
                else:
                    continue

                # Parse amount
                amount_val = row[col_map['amount']]
                if amount_val is None:
                    continue

                if isinstance(amount_val, (int, float)):
                    amount = Decimal(str(amount_val))
                elif isinstance(amount_val, str):
                    clean = amount_val.replace('$', '').replace('.', '').replace(',', '.').strip()
                    try:
                        amount = Decimal(clean)
                    except Exception:
                        errors.append(f'Fila {row_idx}: monto inválido "{amount_val}"')
                        continue
                else:
                    continue

                # Determine transaction type
                tx_type = 'credit' if amount > 0 else 'debit'
                amount = abs(amount)

                # Get description and reference
                description = str(row[col_map.get('description', -1)] or '') if col_map.get('description') is not None else ''
                reference = str(row[col_map.get('reference', -1)] or '') if col_map.get('reference') is not None else ''

                # Check for duplicates
                existing = BankTransaction.objects(
                    tenant=tenant, date=tx_date, amount=amount, description=description[:500]
                ).first()
                if existing:
                    duplicates += 1
                    continue

                # Create transaction
                tx = BankTransaction(
                    date=tx_date,
                    amount=amount,
                    description=description[:500],
                    reference=reference[:100],
                    transaction_type=tx_type,
                    status='pending',
                    source_file=filename,
                    row_number=row_idx,
                    tenant=tenant
                )
                tx.save()
                created += 1

            except Exception as e:
                errors.append(f'Fila {row_idx}: {str(e)}')

        wb.close()

        # Log activity
        ActivityLog.log(
            user=current_user,
            action='create',
            module='reconciliation',
            description=f'Importó {created} transacciones bancarias desde "{filename}"',
            details={'file': filename, 'created': created, 'duplicates': duplicates, 'errors': len(errors)},
            request=request,
            tenant=tenant
        )

        return jsonify({
            'success': True,
            'created': created,
            'duplicates': duplicates,
            'errors': errors[:20],
            'total_errors': len(errors)
        })
        
    except Exception as e:
        return jsonify({'error': f'Error al procesar archivo: {str(e)}'}), 500


@bp.route('/api/transactions/<tx_id>/match', methods=['POST'])
@login_required
@permission_required('reconciliation', 'edit')
def match_transaction(tx_id):
    """Manually match a transaction with a sale"""
    tenant = g.current_tenant
    data = request.get_json()
    
    sale_id = data.get('sale_id')
    if not sale_id:
        return jsonify({'error': 'sale_id es requerido'}), 400
    
    try:
        tx = BankTransaction.objects.get(id=ObjectId(tx_id), tenant=tenant)
    except Exception as e:
        logger.warning(f"match_transaction: Transacción no encontrada {tx_id} - {e}")
        return jsonify({'error': 'Transacción no encontrada'}), 404
    
    try:
        sale = Sale.objects.get(id=ObjectId(sale_id), tenant=tenant)
    except Exception as e:
        logger.warning(f"match_transaction: Venta no encontrada {sale_id} - {e}")
        return jsonify({'error': 'Venta no encontrada'}), 404

    # Check if sale is already matched with another transaction
    existing = BankTransaction.objects(tenant=tenant, matched_sale=sale, status='matched').first()
    if existing:
        return jsonify({'error': 'Esta venta ya está conciliada con otra transacción'}), 400

    # Check amount difference and warn
    sale_total = sale.total_amount
    warning = None
    amount_diff_pct = 0
    if sale_total > 0:
        amount_diff_pct = abs(float(tx.amount) - sale_total) / sale_total * 100
        if amount_diff_pct > 10:
            warning = f'Diferencia de monto: {amount_diff_pct:.1f}% (Transacción: ${float(tx.amount):,.0f} / Venta: ${sale_total:,.0f})'

    # Match
    tx.matched_sale = sale
    tx.status = 'matched'
    tx.match_type = 'manual'
    tx.matched_at = utc_now()
    tx.matched_by = current_user
    tx.save()

    # Create Payment record
    payment = Payment(
        sale=sale,
        tenant=tenant,
        amount=tx.amount,
        payment_via='transferencia',
        payment_reference=f'Conciliación bancaria #{str(tx.id)[-6:]}',
        notes=f'Conciliado desde cartola: {tx.description or ""}',
        created_by=current_user
    )
    payment.save()

    # Recalculate sale payment_status from payments
    sale.payment_status = sale.computed_payment_status
    sale.save()

    # Log activity
    ActivityLog.log(
        user=current_user,
        action='update',
        module='reconciliation',
        description=f'Concilió transacción ${float(tx.amount):,.0f} con venta de "{sale.customer_name}"',
        target_id=str(tx.id),
        target_type='BankTransaction',
        request=request,
        tenant=tenant
    )

    result = {
        'success': True,
        'message': 'Transacción conciliada exitosamente'
    }
    if warning:
        result['warning'] = warning
        result['amount_diff'] = round(amount_diff_pct, 1)
        result['sale_total'] = sale_total

    return jsonify(result)


@bp.route('/api/transactions/<tx_id>/unmatch', methods=['POST'])
@login_required
@permission_required('reconciliation', 'edit')
def unmatch_transaction(tx_id):
    """Remove match from a transaction and revert payment"""
    tenant = g.current_tenant

    try:
        tx = BankTransaction.objects.get(id=ObjectId(tx_id), tenant=tenant)
    except Exception as e:
        logger.warning(f"unmatch_transaction: Transacción no encontrada {tx_id} - {e}")
        return jsonify({'error': 'Transacción no encontrada'}), 404

    # Save sale reference before clearing
    sale = tx.matched_sale

    # Delete the Payment created by this reconciliation
    if sale:
        payment_ref = f'Conciliación bancaria #{str(tx.id)[-6:]}'
        Payment.objects(sale=sale, payment_reference=payment_ref).delete()

    tx.matched_sale = None
    tx.status = 'pending'
    tx.match_type = None
    tx.matched_at = None
    tx.matched_by = None
    tx.save()

    # Recalculate sale payment_status
    if sale:
        sale.payment_status = sale.computed_payment_status
        sale.save()

    ActivityLog.log(
        user=current_user,
        action='update',
        module='reconciliation',
        description=f'Deshizo conciliación de transacción ${float(tx.amount):,.0f}',
        target_id=str(tx.id),
        target_type='BankTransaction',
        request=request,
        tenant=tenant
    )

    return jsonify({
        'success': True,
        'message': 'Conciliación removida'
    })


@bp.route('/api/transactions/<tx_id>/ignore', methods=['POST'])
@login_required
@permission_required('reconciliation', 'edit')
def ignore_transaction(tx_id):
    """Mark transaction as ignored"""
    tenant = g.current_tenant
    
    try:
        tx = BankTransaction.objects.get(id=ObjectId(tx_id), tenant=tenant)
    except Exception as e:
        logger.warning(f"ignore_transaction: Transacción no encontrada {tx_id} - {e}")
        return jsonify({'error': 'Transacción no encontrada'}), 404
    
    tx.status = 'ignored'
    tx.save()

    ActivityLog.log(
        user=current_user,
        action='update',
        module='reconciliation',
        description=f'Ignoró transacción ${float(tx.amount):,.0f} - "{tx.description or ""}"',
        target_id=str(tx.id),
        target_type='BankTransaction',
        request=request,
        tenant=tenant
    )

    return jsonify({
        'success': True,
        'message': 'Transacción ignorada'
    })


@bp.route('/api/transactions/<tx_id>/suggestions', methods=['GET'])
@login_required
@permission_required('reconciliation', 'view')
def get_match_suggestions(tx_id):
    """Get auto-match suggestions for a transaction"""
    tenant = g.current_tenant
    
    try:
        tx = BankTransaction.objects.get(id=ObjectId(tx_id), tenant=tenant)
    except Exception as e:
        logger.warning(f"get_match_suggestions: Transacción no encontrada {tx_id} - {e}")
        return jsonify({'error': 'Transacción no encontrada'}), 404
    
    # Search criteria: amount ±1%, date ±3 days
    amount = float(tx.amount)
    amount_min = amount * 0.99
    amount_max = amount * 1.01
    date_min = tx.date - timedelta(days=3)
    date_max = tx.date + timedelta(days=3)
    
    # Get unmatched sales in range
    suggestions = []
    
    # Get sales that haven't been matched yet and are in the date/amount range
    sales = Sale.objects(
        tenant=tenant,
        date_created__gte=date_min,
        date_created__lte=date_max,
        payment_status__in=['pendiente', 'parcial']  # Only unpaid/partial
    )
    
    for sale in sales:
        sale_total = sale.total_amount
        if amount_min <= sale_total <= amount_max:
            # Calculate confidence score
            amount_diff = abs(sale_total - amount) / amount * 100
            date_diff = abs((sale.date_created - tx.date).days)
            confidence = 100 - (amount_diff * 5) - (date_diff * 10)
            confidence = max(0, min(100, confidence))
            
            suggestions.append({
                'sale_id': str(sale.id),
                'customer': sale.customer_name,
                'total': sale_total,
                'date': sale.date_created.strftime('%Y-%m-%d'),
                'confidence': round(confidence),
                'amount_diff': round(amount_diff, 2),
                'date_diff': date_diff
            })
    
    # Sort by confidence
    suggestions.sort(key=lambda x: x['confidence'], reverse=True)
    
    return jsonify({
        'transaction_id': str(tx.id),
        'transaction_amount': float(tx.amount),
        'suggestions': suggestions[:10]
    })


@bp.route('/api/transactions/auto-match', methods=['POST'])
@login_required
@permission_required('reconciliation', 'edit')
def auto_match_all():
    """Auto-match all pending transactions with high confidence"""
    tenant = g.current_tenant
    
    matched = 0
    errors = []
    
    # Get all pending transactions
    pending_txs = BankTransaction.objects(tenant=tenant, status='pending', transaction_type='credit')
    
    for tx in pending_txs:
        try:
            amount = float(tx.amount)
            amount_min = amount * 0.99
            amount_max = amount * 1.01
            date_min = tx.date - timedelta(days=3)
            date_max = tx.date + timedelta(days=3)
            
            # Find best match
            best_match = None
            best_confidence = 0
            
            sales = Sale.objects(
                tenant=tenant,
                date_created__gte=date_min,
                date_created__lte=date_max,
                payment_status__in=['pendiente', 'parcial']
            )
            
            for sale in sales:
                # Skip if already matched to another transaction
                existing_match = BankTransaction.objects(
                    tenant=tenant,
                    matched_sale=sale,
                    status='matched'
                ).first()
                if existing_match:
                    continue
                
                sale_total = sale.total_amount
                if amount_min <= sale_total <= amount_max:
                    amount_diff = abs(sale_total - amount) / amount * 100
                    date_diff = abs((sale.date_created - tx.date).days)
                    confidence = 100 - (amount_diff * 5) - (date_diff * 10)
                    
                    if confidence >= 80 and confidence > best_confidence:
                        best_match = sale
                        best_confidence = confidence
            
            # Auto-match if confidence >= 80%
            if best_match and best_confidence >= 80:
                tx.matched_sale = best_match
                tx.status = 'matched'
                tx.match_type = 'auto'
                tx.matched_at = utc_now()
                tx.matched_by = current_user
                tx.save()

                # Create Payment record
                payment = Payment(
                    sale=best_match,
                    tenant=tenant,
                    amount=tx.amount,
                    payment_via='transferencia',
                    payment_reference=f'Conciliación bancaria #{str(tx.id)[-6:]}',
                    notes=f'Auto-conciliado desde cartola: {tx.description or ""}',
                    created_by=current_user
                )
                payment.save()

                best_match.payment_status = best_match.computed_payment_status
                best_match.save()

                matched += 1
                
        except Exception as e:
            errors.append(f'TX {tx.id}: {str(e)}')
    
    # Log activity
    ActivityLog.log(
        user=current_user,
        action='update',
        module='reconciliation',
        description=f'Auto-concilió {matched} transacciones',
        details={'matched': matched, 'errors': len(errors)},
        request=request,
        tenant=tenant
    )
    
    return jsonify({
        'success': True,
        'matched': matched,
        'errors': errors[:10]
    })


@bp.route('/api/stats', methods=['GET'])
@login_required
@permission_required('reconciliation', 'view')
def get_stats():
    """Get reconciliation statistics"""
    tenant = g.current_tenant
    
    total = BankTransaction.objects(tenant=tenant).count()
    pending = BankTransaction.objects(tenant=tenant, status='pending').count()
    matched = BankTransaction.objects(tenant=tenant, status='matched').count()
    ignored = BankTransaction.objects(tenant=tenant, status='ignored').count()
    
    # Sum of pending credits (potential income)
    pending_amount = sum(
        float(t.amount) for t in 
        BankTransaction.objects(tenant=tenant, status='pending', transaction_type='credit')
    )
    
    # Sum of matched credits (confirmed income)
    matched_amount = sum(
        float(t.amount) for t in 
        BankTransaction.objects(tenant=tenant, status='matched', transaction_type='credit')
    )
    
    return jsonify({
        'total': total,
        'pending': pending,
        'matched': matched,
        'ignored': ignored,
        'pending_amount': pending_amount,
        'matched_amount': matched_amount,
        'match_rate': round(matched / total * 100, 1) if total > 0 else 0
    })


@bp.route('/api/sales/unmatched', methods=['GET'])
@login_required
@permission_required('reconciliation', 'view')
def get_unmatched_sales():
    """Get sales that haven't been matched with any transaction"""
    tenant = g.current_tenant
    
    # Get all matched sale IDs
    matched_sale_ids = [
        t.matched_sale.id for t in 
        BankTransaction.objects(tenant=tenant, status='matched')
        if t.matched_sale
    ]
    
    # Get unmatched sales
    sales = Sale.objects(
        tenant=tenant,
        id__nin=matched_sale_ids,
        payment_status__in=['pendiente', 'parcial']
    ).order_by('-date_created').limit(100)
    
    results = []
    for s in sales:
        results.append({
            'id': str(s.id),
            'customer': s.customer_name,
            'total': s.total_amount,
            'date': s.date_created.strftime('%Y-%m-%d') if s.date_created else None,
            'payment_status': s.payment_status
        })
    
    return jsonify({'sales': results})


@bp.route('/api/export', methods=['GET'])
@login_required
@permission_required('reconciliation', 'export')
def export_reconciliation():
    """Export reconciliation report to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import io

    tenant = g.current_tenant

    # Apply same filters as the view
    status = request.args.get('status', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    query = BankTransaction.objects(tenant=tenant)

    if status:
        query = query.filter(status=status)
    if date_from:
        try:
            query = query.filter(date__gte=datetime.strptime(date_from, '%Y-%m-%d'))
        except ValueError:
            pass
    if date_to:
        try:
            query = query.filter(date__lt=datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1))
        except ValueError:
            pass

    transactions = query.order_by('-date')

    wb = Workbook()
    ws = wb.active
    ws.title = "Cuadratura Bancaria"

    headers = ['Fecha', 'Descripción', 'Referencia', 'Monto', 'Tipo', 'Estado', 'Venta Asociada', 'Cliente', 'Fecha Match']
    ws.append(headers)

    # Header styles
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Status colors
    matched_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    pending_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    ignored_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    status_labels = {'pending': 'Pendiente', 'matched': 'Conciliada', 'ignored': 'Ignorada'}
    type_labels = {'credit': 'Ingreso', 'debit': 'Egreso'}

    for t in transactions:
        sale_id = ''
        customer = ''
        if t.matched_sale:
            sale_id = str(t.matched_sale.id)[-6:]
            customer = t.matched_sale.customer_name or ''

        row = [
            t.date.strftime('%d/%m/%Y') if t.date else '',
            t.description or '',
            t.reference or '',
            float(t.amount) if t.amount else 0,
            type_labels.get(t.transaction_type, t.transaction_type),
            status_labels.get(t.status, t.status),
            sale_id,
            customer,
            t.matched_at.strftime('%d/%m/%Y %H:%M') if t.matched_at else '',
        ]
        ws.append(row)

        # Apply status color to the row
        row_idx = ws.max_row
        fill = matched_fill if t.status == 'matched' else pending_fill if t.status == 'pending' else ignored_fill
        for cell in ws[row_idx]:
            cell.fill = fill

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 40)

    # Format amount column as number
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            cell.number_format = '#,##0'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'cuadratura_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@bp.route('/api/transactions/batch', methods=['POST'])
@login_required
@permission_required('reconciliation', 'edit')
def batch_transactions():
    """Batch operations on transactions"""
    tenant = g.current_tenant
    data = request.get_json()

    action = data.get('action')
    ids = data.get('ids', [])

    if action != 'ignore':
        return jsonify({'error': 'Acción no soportada. Solo se permite "ignore".'}), 400

    if not ids:
        return jsonify({'error': 'No se proporcionaron IDs'}), 400

    # Validate and process
    processed = 0
    for tx_id in ids:
        try:
            tx = BankTransaction.objects.get(id=ObjectId(tx_id), tenant=tenant, status='pending')
            tx.status = 'ignored'
            tx.save()
            processed += 1
        except Exception:
            continue

    ActivityLog.log(
        user=current_user,
        action='update',
        module='reconciliation',
        description=f'Ignoró {processed} transacciones en lote',
        details={'ids': ids[:10], 'processed': processed},
        request=request,
        tenant=tenant
    )

    return jsonify({
        'success': True,
        'processed': processed
    })
