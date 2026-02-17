import os
import sys
import time
from flask import Blueprint, jsonify, request, g, render_template, send_file
from flask_login import login_required, current_user
from app.models import ShopifyCustomer, ShopifyOrder, ShopifyOrderLineItem, Tenant, utc_now
from datetime import datetime, timedelta
from bson import ObjectId
from functools import wraps
import requests
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
import gspread
from google.oauth2.service_account import Credentials

# Add scripts directory to path for shopify_auth import
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', '..', 'scripts'))
try:
    from shopify_auth import get_auth_headers, get_access_token
    SHOPIFY_AUTH_AVAILABLE = True
except ImportError:
    SHOPIFY_AUTH_AVAILABLE = False

bp = Blueprint('customers', __name__, url_prefix='/customers')

# Shopify API Configuration
SHOPIFY_STORE = os.environ.get('SHOPIFY_STORE_DOMAIN', '')
SHOPIFY_API_VERSION = '2026-01'
SHOPIFY_BASE_URL = f'https://{SHOPIFY_STORE}/admin/api/{SHOPIFY_API_VERSION}'


def get_shopify_headers():
    """Get Shopify API headers with valid access token (auto-refreshes)"""
    if SHOPIFY_AUTH_AVAILABLE:
        try:
            return get_auth_headers()
        except Exception as e:
            raise RuntimeError(f"Error obteniendo token Shopify: {str(e)}")
    else:
        # Fallback to legacy static token
        token = os.environ.get('SHOPIFY_ACCESS_TOKEN', '')
        if not token:
            raise RuntimeError("SHOPIFY_ACCESS_TOKEN no configurado")
        return {
            'X-Shopify-Access-Token': token,
            'Content-Type': 'application/json'
        }


def get_google_sheet():
    """Connect to the ManyChat Google Sheet"""
    creds_file = os.environ.get('GOOGLE_SHEETS_CREDENTIALS_FILE')
    sheet_id = os.environ.get('GOOGLE_SHEETS_ID')
    if not creds_file or not sheet_id:
        raise RuntimeError('GOOGLE_SHEETS_CREDENTIALS_FILE y GOOGLE_SHEETS_ID son requeridos en .env')
    scopes = ['https://www.googleapis.com/auth/spreadsheets.readonly']
    credentials = Credentials.from_service_account_file(creds_file, scopes=scopes)
    client = gspread.authorize(credentials)
    return client.open_by_key(sheet_id).sheet1


def permission_required(module, action='view'):
    """Decorator to check permissions before accessing a route"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                return jsonify({'error': 'No autenticado'}), 401
            if not current_user.has_permission(module, action):
                return jsonify({'error': 'No tienes permisos para esta acción'}), 403
            return f(*args, **kwargs)
        return decorated_function
    return decorator


@bp.route('/')
@login_required
def customers_view():
    """Render customers view page"""
    if not current_user.has_permission('customers', 'view'):
        return "No tienes permisos para ver esta página", 403
    return render_template('customers.html')


@bp.route('/api/customers')
@login_required
@permission_required('customers', 'view')
def get_customers():
    """Get customers list with search and pagination"""
    tenant = g.current_tenant
    
    # Get query parameters
    search = request.args.get('q', '').strip()
    tag_filter = request.args.get('tag', '').strip()
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 50))
    
    # Build query
    if search:
        customers = ShopifyCustomer.objects(
            tenant=tenant
        ).filter(
            __raw__={'$or': [
                {'name': {'$regex': search, '$options': 'i'}},
                {'email': {'$regex': search, '$options': 'i'}},
                {'phone': {'$regex': search, '$options': 'i'}},
            ]}
        ).order_by('-total_spent')
    else:
        customers = ShopifyCustomer.objects(tenant=tenant).order_by('-total_spent')

    if tag_filter:
        customers = customers.filter(tags=tag_filter)

    # Pagination
    total = customers.count()
    customers = customers.skip((page - 1) * per_page).limit(per_page)
    
    # Format results
    results = []
    for c in customers:
        results.append({
            'id': str(c.id),
            'name': c.name or 'Sin nombre',
            'email': c.email or '',
            'phone': c.phone or '',
            'city': c.address_city or '',
            'province': c.address_province or '',
            'country': c.address_country or '',
            'total_orders': c.total_orders or 0,
            'total_spent': float(c.total_spent) if c.total_spent else 0,
            'last_order_date': c.last_order_date.strftime('%Y-%m-%d') if c.last_order_date else None,
            'created_at': c.created_at.strftime('%Y-%m-%d') if c.created_at else None,
            'tags': c.tags or [],
        })
    
    return jsonify({
        'customers': results,
        'total': total,
        'page': page,
        'per_page': per_page,
        'pages': (total + per_page - 1) // per_page
    })


@bp.route('/api/customers/<customer_id>')
@login_required
@permission_required('customers', 'view')
def get_customer_detail(customer_id):
    """Get customer detail with recent orders"""
    tenant = g.current_tenant
    
    try:
        customer = ShopifyCustomer.objects.get(id=ObjectId(customer_id), tenant=tenant)
    except Exception:
        return jsonify({'error': 'Cliente no encontrado'}), 404
    
    # Get recent orders
    orders = ShopifyOrder.objects(customer=customer, tenant=tenant).order_by('-created_at').limit(10)
    
    orders_data = []
    for order in orders:
        orders_data.append({
            'id': str(order.id),
            'order_number': order.order_number,
            'created_at': order.created_at.strftime('%Y-%m-%d %H:%M') if order.created_at else None,
            'total_price': float(order.total_price) if order.total_price else 0,
            'financial_status': order.financial_status,
            'fulfillment_status': order.fulfillment_status or 'unfulfilled',
            'line_items_count': len(order.line_items) if order.line_items else 0,
            'line_items': [
                {
                    'title': item.title,
                    'quantity': item.quantity,
                    'price': float(item.price) if item.price else 0,
                }
                for item in (order.line_items or [])
            ]
        })
    
    return jsonify({
        'customer': {
            'id': str(customer.id),
            'name': customer.name,
            'email': customer.email,
            'phone': customer.phone,
            'city': customer.address_city,
            'province': customer.address_province,
            'country': customer.address_country,
            'total_orders': customer.total_orders,
            'total_spent': float(customer.total_spent) if customer.total_spent else 0,
            'first_order_date': customer.first_order_date.strftime('%Y-%m-%d') if customer.first_order_date else None,
            'last_order_date': customer.last_order_date.strftime('%Y-%m-%d') if customer.last_order_date else None,
            'tags': customer.tags,
        },
        'orders': orders_data
    })


@bp.route('/api/customers', methods=['POST'])
@login_required
@permission_required('customers', 'create')
def create_customer():
    """Create a manual customer"""
    tenant = g.current_tenant
    data = request.get_json()

    if not data or not data.get('name', '').strip():
        return jsonify({'error': 'El nombre es requerido'}), 400

    # Generate unique shopify_id for manual customers
    shopify_id = f"MANUAL-{int(time.time() * 1000)}"

    try:
        customer = ShopifyCustomer(
            name=data['name'].strip(),
            email=data.get('email', '').strip() or None,
            phone=data.get('phone', '').strip() or None,
            address_city=data.get('address_city', '').strip() or None,
            address_province=data.get('address_province', '').strip() or None,
            address_country=data.get('address_country', '').strip() or None,
            shopify_id=shopify_id,
            source='manual',
            total_orders=0,
            total_spent=0,
            created_at=utc_now(),
            updated_at=utc_now(),
            tenant=tenant
        )
        customer.save()

        return jsonify({
            'success': True,
            'message': 'Cliente creado exitosamente',
            'customer': {
                'id': str(customer.id),
                'name': customer.name,
                'email': customer.email or '',
                'source': customer.source,
            }
        }), 201
    except Exception as e:
        return jsonify({'error': f'Error al crear cliente: {str(e)}'}), 500


@bp.route('/api/customers/import', methods=['POST'])
@login_required
@permission_required('customers', 'sync')
def import_customers():
    """Import customers from Excel file"""
    tenant = g.current_tenant

    if 'file' not in request.files:
        return jsonify({'error': 'No se proporcionó archivo'}), 400

    file = request.files['file']
    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': 'El archivo debe ser .xlsx'}), 400

    try:
        wb = load_workbook(file, read_only=True)
        ws = wb.active

        # Read header row
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        # Map column names (case-insensitive)
        col_map = {}
        field_aliases = {
            'nombre': 'name',
            'name': 'name',
            'email': 'email',
            'correo': 'email',
            'teléfono': 'phone',
            'telefono': 'phone',
            'phone': 'phone',
            'ciudad': 'city',
            'city': 'city',
            'provincia': 'province',
            'province': 'province',
            'región': 'province',
            'region': 'province',
            'país': 'country',
            'pais': 'country',
            'country': 'country',
        }
        for idx, h in enumerate(headers):
            if h:
                key = h.strip().lower()
                if key in field_aliases:
                    col_map[field_aliases[key]] = idx

        if 'name' not in col_map:
            return jsonify({'error': 'La columna "Nombre" es requerida en el archivo Excel'}), 400

        # Check if this is a preview request
        preview = request.args.get('preview', '').lower() == 'true'

        rows_data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = row[col_map['name']] if col_map.get('name') is not None and row[col_map['name']] else None
            if not name:
                continue
            rows_data.append({
                'name': str(name).strip(),
                'email': str(row[col_map['email']]).strip() if col_map.get('email') is not None and row[col_map['email']] else '',
                'phone': str(row[col_map['phone']]).strip() if col_map.get('phone') is not None and row[col_map['phone']] else '',
                'city': str(row[col_map['city']]).strip() if col_map.get('city') is not None and row[col_map['city']] else '',
                'province': str(row[col_map['province']]).strip() if col_map.get('province') is not None and row[col_map['province']] else '',
                'country': str(row[col_map['country']]).strip() if col_map.get('country') is not None and row[col_map['country']] else '',
            })

        wb.close()

        if preview:
            return jsonify({
                'preview': True,
                'total': len(rows_data),
                'rows': rows_data[:100],  # Preview first 100
            })

        # Import all rows
        timestamp = int(time.time())
        created = 0
        errors = []

        for idx, row_data in enumerate(rows_data):
            try:
                shopify_id = f"IMPORT-{idx + 1}-{timestamp}"
                customer = ShopifyCustomer(
                    name=row_data['name'],
                    email=row_data['email'] or None,
                    phone=row_data['phone'] or None,
                    address_city=row_data['city'] or None,
                    address_province=row_data['province'] or None,
                    address_country=row_data['country'] or None,
                    shopify_id=shopify_id,
                    source='import',
                    total_orders=0,
                    total_spent=0,
                    created_at=utc_now(),
                    updated_at=utc_now(),
                    tenant=tenant
                )
                customer.save()
                created += 1
            except Exception as e:
                errors.append(f'Fila {idx + 2}: {str(e)}')

        return jsonify({
            'success': True,
            'created': created,
            'total': len(rows_data),
            'errors': errors[:10],
        })

    except Exception as e:
        return jsonify({'error': f'Error al procesar archivo: {str(e)}'}), 500


@bp.route('/api/customers/stats')
@login_required
@permission_required('customers', 'view')
def get_stats():
    """Get customer statistics"""
    tenant = g.current_tenant
    
    # Total customers
    total_customers = ShopifyCustomer.objects(tenant=tenant).count()
    
    # Total revenue
    from decimal import Decimal
    total_revenue = Decimal('0')
    for customer in ShopifyCustomer.objects(tenant=tenant).only('total_spent'):
        if customer.total_spent:
            total_revenue += customer.total_spent
    
    # Average ticket
    avg_ticket = float(total_revenue / total_customers) if total_customers > 0 else 0
    
    # New customers this month
    start_of_month = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    new_this_month = ShopifyCustomer.objects(tenant=tenant, created_at__gte=start_of_month).count()
    
    return jsonify({
        'total_customers': total_customers,
        'total_revenue': float(total_revenue),
        'avg_ticket': avg_ticket,
        'new_this_month': new_this_month,
    })


@bp.route('/api/customers/export')
@login_required
@permission_required('customers', 'export')
def export_excel():
    """Export customers to Excel"""
    tenant = g.current_tenant
    
    # Get all customers
    customers = ShopifyCustomer.objects(tenant=tenant).order_by('-total_spent')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes Shopify"
    
    # Header style
    header_fill = PatternFill(start_color="C85103", end_color="C85103", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = ['Nombre', 'Email', 'Teléfono', 'Ciudad', 'Provincia', 'País', 
               'Total Pedidos', 'Total Gastado', 'Primer Pedido', 'Último Pedido']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Data rows
    for row_idx, customer in enumerate(customers, start=2):
        ws.cell(row=row_idx, column=1, value=customer.name or '')
        ws.cell(row=row_idx, column=2, value=customer.email or '')
        ws.cell(row=row_idx, column=3, value=customer.phone or '')
        ws.cell(row=row_idx, column=4, value=customer.address_city or '')
        ws.cell(row=row_idx, column=5, value=customer.address_province or '')
        ws.cell(row=row_idx, column=6, value=customer.address_country or '')
        ws.cell(row=row_idx, column=7, value=customer.total_orders or 0)
        ws.cell(row=row_idx, column=8, value=float(customer.total_spent) if customer.total_spent else 0)
        ws.cell(row=row_idx, column=9, value=customer.first_order_date.strftime('%Y-%m-%d') if customer.first_order_date else '')
        ws.cell(row=row_idx, column=10, value=customer.last_order_date.strftime('%Y-%m-%d') if customer.last_order_date else '')
    
    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 15
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'clientes_shopify_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )


@bp.route('/api/customers', methods=['POST'])
@login_required
@permission_required('customers', 'create')
def api_create_customer_v2():
    """Create a manual customer"""
    tenant = g.current_tenant
    data = request.get_json()
    
    if not data or not data.get('name', '').strip():
        return jsonify({'error': 'El nombre es requerido'}), 400
    
    import time
    customer = ShopifyCustomer(
        name=data['name'].strip(),
        email=data.get('email', '').strip() or None,
        phone=data.get('phone', '').strip() or None,
        address_city=data.get('address_city', '').strip() or None,
        address_province=data.get('address_province', '').strip() or None,
        address_country=data.get('address_country', '').strip() or 'Chile',
        source='manual',
        shopify_id=f"MANUAL-{int(time.time() * 1000)}",
        created_at=utc_now(),
        updated_at=utc_now(),
        tenant=tenant
    )
    customer.save()
    
    return jsonify({
        'id': str(customer.id),
        'name': customer.name,
        'message': 'Cliente creado exitosamente'
    }), 201


@bp.route('/api/customers/import', methods=['POST'])
@login_required
@permission_required('customers', 'sync')
def api_import_customers_v2():
    """Import customers from Excel file"""
    tenant = g.current_tenant
    
    if 'file' not in request.files:
        return jsonify({'error': 'No se envió archivo'}), 400
    
    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Solo se aceptan archivos Excel (.xlsx)'}), 400
    
    try:
        from openpyxl import load_workbook
        import time
        
        wb = load_workbook(file, read_only=True)
        ws = wb.active
        
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        
        imported = 0
        errors = []
        timestamp = int(time.time())
        
        for idx, row in enumerate(rows):
            try:
                name = str(row[0]).strip() if row[0] else None
                if not name or name == 'None':
                    continue
                
                email = str(row[1]).strip() if len(row) > 1 and row[1] else None
                phone = str(row[2]).strip() if len(row) > 2 and row[2] else None
                city = str(row[3]).strip() if len(row) > 3 and row[3] else None
                province = str(row[4]).strip() if len(row) > 4 and row[4] else None
                country = str(row[5]).strip() if len(row) > 5 and row[5] else 'Chile'
                
                if email and email != 'None':
                    existing = ShopifyCustomer.objects(email=email, tenant=tenant).first()
                    if existing:
                        errors.append(f'Fila {idx+2}: {name} ya existe (email duplicado)')
                        continue
                
                customer = ShopifyCustomer(
                    name=name,
                    email=email if email and email != 'None' else None,
                    phone=phone if phone and phone != 'None' else None,
                    address_city=city if city and city != 'None' else None,
                    address_province=province if province and province != 'None' else None,
                    address_country=country if country and country != 'None' else 'Chile',
                    source='import',
                    shopify_id=f"IMPORT-{idx}-{timestamp}",
                    created_at=utc_now(),
                    updated_at=utc_now(),
                    tenant=tenant
                )
                customer.save()
                imported += 1
                
            except Exception as e:
                errors.append(f'Fila {idx+2}: {str(e)}')
        
        wb.close()
        
        return jsonify({
            'imported': imported,
            'total_rows': len(rows),
            'errors': errors
        })
        
    except Exception as e:
        return jsonify({'error': f'Error al procesar archivo: {str(e)}'}), 500


@bp.route('/api/customers/<customer_id>/tags', methods=['PUT'])
@login_required
@permission_required('customers', 'edit')
def update_customer_tags(customer_id):
    """Update customer tags"""
    tenant = g.current_tenant
    data = request.get_json()

    if not data or 'tags' not in data:
        return jsonify({'error': 'Tags requeridos'}), 400

    try:
        customer = ShopifyCustomer.objects.get(id=ObjectId(customer_id), tenant=tenant)
    except Exception:
        return jsonify({'error': 'Cliente no encontrado'}), 404

    customer.tags = [t.strip().lower() for t in data['tags'] if t.strip()]
    customer.updated_at = utc_now()
    customer.save()

    return jsonify({'success': True, 'tags': customer.tags})


@bp.route('/api/customers/sync', methods=['POST'])
@login_required
@permission_required('customers', 'sync')
def sync_shopify():
    """Sync customers and orders from Shopify (admin only)"""
    tenant = g.current_tenant
    
    try:
        headers = get_shopify_headers()
    except RuntimeError as e:
        return jsonify({'error': str(e)}), 500
    
    stats = {
        'customers_synced': 0,
        'orders_synced': 0,
        'errors': []
    }
    
    # Flag para saber si debemos extraer clientes de órdenes
    extract_customers_from_orders = False
    
    try:
        # Intentar Sync Customers (puede fallar si no hay scope read_customers)
        customers_url = f'{SHOPIFY_BASE_URL}/customers.json'
        response = requests.get(customers_url, headers=headers, params={'limit': 1})
        
        if response.status_code == 403:
            # No tenemos permiso para leer clientes directamente
            # Los extraeremos de las órdenes
            extract_customers_from_orders = True
            stats['errors'].append('Sin permiso read_customers - extrayendo clientes desde órdenes')
        elif response.status_code == 200:
            # Sí tenemos acceso, sincronizar normalmente
            page_info = None
            
            while True:
                params = {'limit': 250}
                if page_info:
                    params['page_info'] = page_info
                
                response = requests.get(customers_url, headers=headers, params=params)
                
                if response.status_code != 200:
                    stats['errors'].append(f'Error al obtener clientes: {response.status_code}')
                    break
                
                data = response.json()
                customers_data = data.get('customers', [])
                
                if not customers_data:
                    break
                
                # Process customers
                for customer_data in customers_data:
                    try:
                        shopify_id = str(customer_data['id'])
                        
                        # Get default address
                        default_address = customer_data.get('default_address') or {}
                        
                        # Upsert customer
                        customer = ShopifyCustomer.objects(shopify_id=shopify_id, tenant=tenant).first()
                        if not customer:
                            customer = ShopifyCustomer(shopify_id=shopify_id, tenant=tenant)
                        
                        customer.name = f"{customer_data.get('first_name', '')} {customer_data.get('last_name', '')}".strip()
                        customer.email = customer_data.get('email')
                        customer.phone = customer_data.get('phone') or default_address.get('phone')
                        customer.address_city = default_address.get('city')
                        customer.address_province = default_address.get('province')
                        customer.address_country = default_address.get('country')
                        raw_tags = customer_data.get('tags', '')
                        customer.tags = [t.strip().lower() for t in raw_tags.split(',') if t.strip()] if raw_tags else []
                        customer.total_orders = customer_data.get('orders_count', 0)
                        customer.total_spent = float(customer_data.get('total_spent', 0))
                        
                        # Parse dates
                        if customer_data.get('created_at'):
                            customer.created_at = datetime.fromisoformat(customer_data['created_at'].replace('Z', '+00:00'))
                        
                        customer.save()
                        stats['customers_synced'] += 1
                        
                    except Exception as e:
                        stats['errors'].append(f'Error al procesar cliente {customer_data.get("id")}: {str(e)}')
                
                # Check for next page
                link_header = response.headers.get('Link', '')
                if 'rel="next"' in link_header:
                    import re
                    match = re.search(r'page_info=([^&>]+)', link_header)
                    if match:
                        page_info = match.group(1)
                    else:
                        break
                else:
                    break
        
        # Sync Orders
        orders_url = f'{SHOPIFY_BASE_URL}/orders.json'
        page_info = None
        
        while True:
            params = {'limit': 250, 'status': 'any'}
            if page_info:
                params['page_info'] = page_info
            
            response = requests.get(orders_url, headers=headers, params=params)
            
            if response.status_code != 200:
                stats['errors'].append(f'Error al obtener órdenes: {response.status_code}')
                break
            
            data = response.json()
            orders_data = data.get('orders', [])
            
            if not orders_data:
                break
            
            # Process orders
            for order_data in orders_data:
                try:
                    shopify_id = str(order_data['id'])
                    
                    # Find or create customer from order data
                    customer = None
                    if order_data.get('customer'):
                        customer_shopify_id = str(order_data['customer']['id'])
                        customer = ShopifyCustomer.objects(shopify_id=customer_shopify_id, tenant=tenant).first()
                        
                        # Si no existe y debemos extraer de órdenes, crear el cliente
                        if not customer and extract_customers_from_orders:
                            customer_data = order_data['customer']
                            shipping_addr = order_data.get('shipping_address') or {}
                            
                            customer = ShopifyCustomer(
                                shopify_id=customer_shopify_id,
                                tenant=tenant,
                                source='shopify',
                                name=f"{customer_data.get('first_name', '')} {customer_data.get('last_name', '')}".strip() or 'Cliente Shopify',
                                email=customer_data.get('email') or order_data.get('email'),
                                phone=shipping_addr.get('phone') or customer_data.get('phone'),
                                address_city=shipping_addr.get('city'),
                                address_province=shipping_addr.get('province'),
                                address_country=shipping_addr.get('country', 'Chile'),
                                created_at=utc_now(),
                                updated_at=utc_now()
                            )
                            customer.save()
                            stats['customers_synced'] += 1
                    
                    # Upsert order
                    order = ShopifyOrder.objects(shopify_id=shopify_id, tenant=tenant).first()
                    if not order:
                        order = ShopifyOrder(shopify_id=shopify_id, tenant=tenant)
                    
                    order.order_number = order_data.get('order_number')
                    order.customer = customer
                    order.customer_name = order_data.get('customer', {}).get('first_name', '') + ' ' + order_data.get('customer', {}).get('last_name', '')
                    order.email = order_data.get('email')
                    order.total_price = float(order_data.get('total_price', 0))
                    order.subtotal_price = float(order_data.get('subtotal_price', 0))
                    order.financial_status = order_data.get('financial_status')
                    order.fulfillment_status = order_data.get('fulfillment_status')
                    
                    # Shipping address (captura completa)
                    shipping_address = order_data.get('shipping_address') or {}
                    order.shipping_address1 = shipping_address.get('address1')
                    order.shipping_address2 = shipping_address.get('address2')
                    order.shipping_city = shipping_address.get('city')
                    order.shipping_province = shipping_address.get('province')
                    order.shipping_phone = shipping_address.get('phone')
                    order.note = order_data.get('note')
                    
                    # Line items
                    line_items = []
                    for item_data in order_data.get('line_items', []):
                        line_item = ShopifyOrderLineItem(
                            title=item_data.get('title'),
                            sku=item_data.get('sku'),
                            quantity=item_data.get('quantity', 1),
                            price=float(item_data.get('price', 0)),
                            variant_title=item_data.get('variant_title'),
                            product_shopify_id=str(item_data.get('product_id')) if item_data.get('product_id') else None
                        )
                        line_items.append(line_item)
                    order.line_items = line_items
                    
                    # Parse dates
                    if order_data.get('created_at'):
                        order.created_at = datetime.fromisoformat(order_data['created_at'].replace('Z', '+00:00'))
                    
                    order.save()
                    stats['orders_synced'] += 1
                    
                    # Update customer stats
                    if customer:
                        customer_orders = ShopifyOrder.objects(customer=customer, tenant=tenant)
                        customer.total_orders = customer_orders.count()
                        
                        # Calculate total spent and dates
                        total_spent = sum(float(o.total_price or 0) for o in customer_orders)
                        customer.total_spent = total_spent
                        
                        # Get first and last order dates
                        orders_by_date = customer_orders.order_by('created_at')
                        if orders_by_date:
                            customer.first_order_date = orders_by_date.first().created_at
                            customer.last_order_date = orders_by_date.order_by('-created_at').first().created_at
                        
                        customer.save()
                    
                except Exception as e:
                    stats['errors'].append(f'Error al procesar orden {order_data.get("id")}: {str(e)}')
            
            # Check for next page
            link_header = response.headers.get('Link', '')
            if 'rel="next"' in link_header:
                import re
                match = re.search(r'page_info=([^&>]+)', link_header)
                if match:
                    page_info = match.group(1)
                else:
                    break
            else:
                break
        
    except Exception as e:
        stats['errors'].append(f'Error general en clientes/órdenes: {str(e)}')
    
    # ==========================================
    # SYNC PRODUCTS + STOCK
    # ==========================================
    try:
        from app.models import Product, Lot, InboundOrder
        import re as re_mod
        from decimal import Decimal
        
        products_url = f'{SHOPIFY_BASE_URL}/products.json'
        resp = requests.get(products_url, headers=headers, params={'limit': 250})
        
        if resp.status_code == 200:
            products_data = resp.json().get('products', [])
            products_created = 0
            products_updated = 0
            
            for p_data in products_data:
                shopify_id = str(p_data['id'])
                variants = p_data.get('variants', [])
                variant = variants[0] if variants else {}
                sku = variant.get('sku', '')
                price = Decimal(str(variant.get('price', '0')))
                inv_qty = variant.get('inventory_quantity', 0)
                
                # Strip HTML
                desc_html = p_data.get('body_html', '') or ''
                description = re_mod.sub(r'<[^>]+>', '', desc_html).strip()
                
                # Find or create product
                existing = None
                if sku:
                    existing = Product.objects(sku=sku, tenant=tenant).first()
                if not existing:
                    existing = Product.objects(shopify_id=shopify_id, tenant=tenant).first()
                
                if existing:
                    existing.name = p_data.get('title', existing.name)
                    existing.base_price = price
                    existing.shopify_id = shopify_id
                    if description:
                        existing.description = description[:500]
                    existing.save()
                    products_updated += 1
                    product = existing
                else:
                    product = Product(
                        name=p_data.get('title', 'Sin nombre'),
                        sku=sku or f"SHP-{shopify_id[-6:]}",
                        base_price=price,
                        description=description[:500] if description else '',
                        category=p_data.get('product_type', 'Shopify'),
                        shopify_id=shopify_id,
                        critical_stock=10,
                        tenant=tenant
                    )
                    product.save()
                    products_created += 1
                
                # Update stock via Lot
                if inv_qty > 0:
                    lot = Lot.objects(product=product, lot_code=f'SHOPIFY-{product.sku}').first()
                    if lot:
                        lot.quantity_current = inv_qty
                        lot.quantity_initial = max(lot.quantity_initial, inv_qty)
                        lot.save()
                    else:
                        # Create inbound order if needed
                        today_str = utc_now().strftime('%Y%m%d')
                        order = InboundOrder.objects(
                            invoice_number=f'SHOPIFY-SYNC-{today_str}',
                            tenant=tenant
                        ).first()
                        if not order:
                            order = InboundOrder(
                                supplier_name='Sync Shopify',
                                invoice_number=f'SHOPIFY-SYNC-{today_str}',
                                status='received',
                                date_received=utc_now(),
                                notes='Stock sincronizado desde Shopify',
                                tenant=tenant,
                                created_at=utc_now()
                            )
                            order.save()
                        
                        lot = Lot(
                            product=product,
                            order=order,
                            tenant=tenant,
                            lot_code=f'SHOPIFY-{product.sku}',
                            quantity_initial=inv_qty,
                            quantity_current=inv_qty
                        )
                        lot.save()
            
            stats['products_created'] = products_created
            stats['products_updated'] = products_updated
    
    except Exception as e:
        stats['errors'].append(f'Error en sync productos: {str(e)}')
    
    # ==========================================
    # SYNC ORDERS → SALES
    # ==========================================
    try:
        from app.models import Sale, SaleItem
        
        sales_created = 0
        shopify_orders_all = ShopifyOrder.objects(tenant=tenant)
        
        for s_order in shopify_orders_all:
            existing_sale = Sale.objects(tenant=tenant, shopify_order_id=str(s_order.shopify_id)).first()
            if existing_sale:
                continue
            
            addr_parts = [s_order.shipping_city or '', s_order.shipping_province or '']
            address = ', '.join(p for p in addr_parts if p)
            
            new_sale = Sale(
                customer_name=s_order.customer_name or 'Cliente Shopify',
                address=address,
                sale_type='con_despacho',
                sales_channel='shopify',  # NEW: mark as Shopify origin
                delivery_status='entregado' if s_order.fulfillment_status == 'fulfilled' else 'pendiente',
                payment_status='pagado' if s_order.financial_status == 'paid' else 'pendiente',
                date_created=s_order.created_at or utc_now(),
                shopify_order_id=str(s_order.shopify_id),
                shopify_order_number=s_order.order_number,
                tenant=tenant
            )
            new_sale.save()
            
            for item in s_order.line_items:
                product = None
                if item.sku:
                    product = Product.objects(sku=item.sku, tenant=tenant).first()
                if not product and item.product_shopify_id:
                    product = Product.objects(shopify_id=item.product_shopify_id, tenant=tenant).first()
                
                if product:
                    sale_item = SaleItem(
                        sale=new_sale,
                        product=product,
                        quantity=item.quantity or 1,
                        unit_price=float(item.price) if item.price else 0
                    )
                    sale_item.save()
            
            sales_created += 1
        
        stats['sales_created'] = sales_created
    
    except Exception as e:
        stats['errors'].append(f'Error en sync ventas: {str(e)}')
    
    return jsonify(stats)


@bp.route('/api/customers/sync/preview', methods=['GET'])
@login_required
@permission_required('customers', 'sync')
def sync_shopify_preview():
    """
    Preview Shopify sync changes without applying them.
    Returns what WOULD be created/updated if sync is executed.
    """
    tenant = g.current_tenant
    
    try:
        headers = get_shopify_headers()
    except RuntimeError as e:
        return jsonify({'error': str(e), 'errors': [str(e)]}), 500
    
    preview = {
        'products': {'new': [], 'update': [], 'unchanged': 0},
        'customers': {'new': [], 'update': [], 'unchanged': 0},
        'orders': {'new': [], 'unchanged': 0},
        'errors': []
    }
    
    try:
        # ==========================================
        # PREVIEW PRODUCTS
        # ==========================================
        from app.models import Product
        from decimal import Decimal
        import re as re_mod
        
        products_url = f'{SHOPIFY_BASE_URL}/products.json'
        resp = requests.get(products_url, headers=headers, params={'limit': 250})
        
        if resp.status_code == 200:
            products_data = resp.json().get('products', [])
            
            for p_data in products_data:
                shopify_id = str(p_data['id'])
                variants = p_data.get('variants', [])
                variant = variants[0] if variants else {}
                sku = variant.get('sku', '')
                price = float(variant.get('price', '0'))
                inv_qty = variant.get('inventory_quantity', 0)
                name = p_data.get('title', 'Sin nombre')
                
                # Find existing product
                existing = None
                if sku:
                    existing = Product.objects(sku=sku, tenant=tenant).first()
                if not existing:
                    existing = Product.objects(shopify_id=shopify_id, tenant=tenant).first()
                
                if existing:
                    # Check if there are changes
                    changes = []
                    if existing.name != name:
                        changes.append(f'nombre: {existing.name} → {name}')
                    if float(existing.base_price or 0) != price:
                        changes.append(f'precio: ${existing.base_price} → ${price}')
                    if existing.total_stock != inv_qty:
                        changes.append(f'stock: {existing.total_stock} → {inv_qty}')
                    
                    if changes:
                        preview['products']['update'].append({
                            'sku': sku or f"SHP-{shopify_id[-6:]}",
                            'name': name,
                            'price': price,
                            'stock': inv_qty,
                            'changes': changes
                        })
                    else:
                        preview['products']['unchanged'] += 1
                else:
                    preview['products']['new'].append({
                        'sku': sku or f"SHP-{shopify_id[-6:]}",
                        'name': name,
                        'price': price,
                        'stock': inv_qty
                    })
        else:
            preview['errors'].append(f'Error al obtener productos: {resp.status_code}')
        
        # ==========================================
        # PREVIEW CUSTOMERS
        # ==========================================
        customers_url = f'{SHOPIFY_BASE_URL}/customers.json'
        resp = requests.get(customers_url, headers=headers, params={'limit': 250})
        
        if resp.status_code == 200:
            customers_data = resp.json().get('customers', [])
            
            for c_data in customers_data:
                shopify_id = str(c_data['id'])
                name = f"{c_data.get('first_name', '')} {c_data.get('last_name', '')}".strip()
                email = c_data.get('email', '')
                
                existing = ShopifyCustomer.objects(shopify_id=shopify_id, tenant=tenant).first()
                
                if existing:
                    changes = []
                    if existing.name != name:
                        changes.append(f'nombre')
                    if existing.email != email:
                        changes.append(f'email')
                    
                    if changes:
                        preview['customers']['update'].append({
                            'name': name,
                            'email': email,
                            'changes': changes
                        })
                    else:
                        preview['customers']['unchanged'] += 1
                else:
                    preview['customers']['new'].append({
                        'name': name,
                        'email': email
                    })
        else:
            preview['errors'].append(f'Error al obtener clientes: {resp.status_code}')
        
        # ==========================================
        # PREVIEW ORDERS
        # ==========================================
        from app.models import Sale
        
        orders_url = f'{SHOPIFY_BASE_URL}/orders.json'
        resp = requests.get(orders_url, headers=headers, params={'limit': 250, 'status': 'any'})
        
        if resp.status_code == 200:
            orders_data = resp.json().get('orders', [])
            
            for o_data in orders_data:
                shopify_id = str(o_data['id'])
                order_number = o_data.get('order_number')
                
                # Check if already synced to Sale
                existing_sale = Sale.objects(shopify_order_id=shopify_id, tenant=tenant).first()
                existing_order = ShopifyOrder.objects(shopify_id=shopify_id, tenant=tenant).first()
                
                if not existing_sale:
                    preview['orders']['new'].append({
                        'order_number': order_number,
                        'customer': o_data.get('customer', {}).get('first_name', '') + ' ' + o_data.get('customer', {}).get('last_name', ''),
                        'total': float(o_data.get('total_price', 0)),
                        'status': o_data.get('financial_status')
                    })
                else:
                    preview['orders']['unchanged'] += 1
        else:
            preview['errors'].append(f'Error al obtener órdenes: {resp.status_code}')
        
    except Exception as e:
        preview['errors'].append(f'Error general: {str(e)}')
    
    # Calculate summary
    preview['summary'] = {
        'products_new': len(preview['products']['new']),
        'products_update': len(preview['products']['update']),
        'products_unchanged': preview['products']['unchanged'],
        'customers_new': len(preview['customers']['new']),
        'customers_update': len(preview['customers']['update']),
        'customers_unchanged': preview['customers']['unchanged'],
        'orders_new': len(preview['orders']['new']),
        'orders_unchanged': preview['orders']['unchanged'],
        'has_changes': (
            len(preview['products']['new']) > 0 or
            len(preview['products']['update']) > 0 or
            len(preview['customers']['new']) > 0 or
            len(preview['customers']['update']) > 0 or
            len(preview['orders']['new']) > 0
        )
    }
    
    return jsonify(preview)


@bp.route('/api/customers/sync-manychat', methods=['POST'])
@login_required
@permission_required('customers', 'sync')
def sync_manychat():
    """Import leads from ManyChat Google Sheet"""
    tenant = g.current_tenant

    try:
        sheet = get_google_sheet()
    except Exception as e:
        return jsonify({'error': f'Error conectando a Google Sheets: {str(e)}'}), 500

    try:
        records = sheet.get_all_records()
    except Exception as e:
        return jsonify({'error': f'Error leyendo Sheet: {str(e)}'}), 500

    stats = {'created': 0, 'skipped': 0, 'sales_created': 0, 'errors': []}

    from app.models import Sale, SaleItem, Product
    import re

    for idx, row in enumerate(records):
        try:
            phone = str(row.get('User ID', '')).strip()
            name = str(row.get('Nombre', '')).strip()
            semaforo_raw = str(row.get('Semáforo', row.get('Semaforo', ''))).strip()
            city = str(row.get('Ciudad', '')).strip()
            productos_raw = str(row.get('Productos interes', '')).strip()
            lugar_entrega = str(row.get('Lugar Entrega', '')).strip()
            hora_entrega = str(row.get('Hora Entrega estimada', '') or row.get('Hora Entrega', '')).strip()
            metodo_pago = str(row.get('Método de Pago', row.get('Metodo de Pago', ''))).strip()

            if not phone or not name:
                continue

            # Normalize phone
            phone_clean = phone.replace('+', '').replace(' ', '').replace('-', '')

            # Check if already imported (dedup by phone)
            existing = ShopifyCustomer.objects(
                __raw__={'$or': [
                    {'phone': phone},
                    {'phone': phone_clean},
                    {'phone': f'+{phone_clean}'},
                    {'shopify_id': f'MANYCHAT-{phone_clean}'}
                ]},
                tenant=tenant
            ).first()
            if existing:
                stats['skipped'] += 1
                continue

            # Determine tag from semaforo
            semaforo = semaforo_raw.lower()
            if 'calificado' in semaforo:
                tag = 'calificado'
            elif 'interesado' in semaforo:
                tag = 'interesado'
            else:
                tag = 'poco-interesado'

            # Create customer
            customer = ShopifyCustomer(
                name=name,
                phone=phone_clean,
                address_city=city or None,
                source='manychat',
                shopify_id=f'MANYCHAT-{phone_clean}',
                tags=[tag],
                total_orders=0,
                total_spent=0,
                created_at=utc_now(),
                updated_at=utc_now(),
                tenant=tenant
            )
            customer.save()
            stats['created'] += 1

            # If calificado AND has products, create pending sale
            if tag == 'calificado' and productos_raw:
                try:
                    sale = Sale(
                        customer_name=name,
                        address=lugar_entrega or city or '',
                        phone=phone_clean,
                        sale_type='con_despacho',
                        sales_channel='whatsapp',
                        delivery_status='pendiente',
                        payment_status='pendiente',
                        payment_method=metodo_pago.lower() if metodo_pago else None,
                        date_created=utc_now(),
                        tenant=tenant
                    )
                    sale.save()

                    # Parse products: "Promo jurel x2 | Caja Mensual x1"
                    product_entries = [p.strip() for p in productos_raw.split('|') if p.strip()]
                    notes_lines = []

                    for entry in product_entries:
                        match = re.match(r'(.+?)\s*x\s*(\d+)$', entry.strip(), re.IGNORECASE)
                        if match:
                            prod_name = match.group(1).strip()
                            qty = int(match.group(2))
                        else:
                            prod_name = entry.strip()
                            qty = 1

                        # Search product by name
                        product = Product.objects(tenant=tenant, name__icontains=prod_name).first()

                        if product:
                            SaleItem(
                                sale=sale,
                                product=product,
                                quantity=qty,
                                unit_price=float(product.base_price) if product.base_price else 0
                            ).save()
                        else:
                            notes_lines.append(f'Producto no encontrado: {entry}')

                    if notes_lines:
                        sale.delivery_observations = '\n'.join(notes_lines)
                        sale.save()

                    stats['sales_created'] += 1
                except Exception as e:
                    stats['errors'].append(f'Error creando venta para {name}: {str(e)}')

        except Exception as e:
            stats['errors'].append(f'Fila {idx + 2}: {str(e)}')

    return jsonify(stats)
